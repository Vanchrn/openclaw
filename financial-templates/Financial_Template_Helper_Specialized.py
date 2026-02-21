#!/usr/bin/env python3
"""
Specialized Financial Excel Template Generator
Creates industry-specific and specialized financial templates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_retail_pandl_template():
    """Create Retail P&L template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retail P&L"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "RETAIL PROFIT AND LOSS STATEMENT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Sales Section
    row = 3
    ws[f'A{row}'] = "SALES"
    ws[f'A{row}'].font = section_font
    row += 1

    sales_items = [
        ("Gross Sales", 500000.00),
        ("Less: Sales Returns", -15000.00),
        ("Less: Allowances", -5000.00),
        ("Less: Discounts", -10000.00),
        ("NET SALES", 470000.00),
    ]

    for item, value in sales_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "NET" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # COGS Section
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = section_font
    row += 1

    cogs_items = [
        ("Beginning Inventory", 80000.00),
        ("+ Purchases", 280000.00),
        ("+ Freight In", 5000.00),
        ("Cost of Goods Available", 365000.00),
        ("- Ending Inventory", -75000.00),
        ("COST OF GOODS SOLD", 290000.00),
    ]

    for item, value in cogs_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "COST OF GOODS" in item and not "Available" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C8-C14"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C16/C8"
    ws[f'E{row}'].number_format = percent_format
    row += 2

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = section_font
    row += 1

    op_expenses = [
        ("Store Rent", 25000.00),
        ("Utilities", 5000.00),
        ("Salaries & Wages", 80000.00),
        ("Advertising", 15000.00),
        ("Insurance", 8000.00),
        ("Depreciation", 10000.00),
        ("Other Expenses", 7000.00),
        ("TOTAL OPERATING EXPENSES", 150000.00),
    ]

    for item, value in op_expenses:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Operating Income
    ws[f'A{row}'] = "OPERATING INCOME"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C16-C24"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Other Income/Expense
    ws[f'A{row}'] = "OTHER INCOME/EXPENSE"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Interest Expense"
    ws[f'C{row}'] = -5000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C26+C27"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C28/C8"
    ws[f'E{row}'].number_format = percent_format

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 15

    return wb

def create_service_business_pandl_template():
    """Create Service Business P&L template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Service P&L"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "SERVICE BUSINESS PROFIT AND LOSS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Revenue Section
    row = 3
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = section_font
    row += 1

    revenue_items = [
        ("Consulting Services", 150000.00),
        ("Project Services", 100000.00),
        ("Maintenance Contracts", 50000.00),
        ("Training Services", 30000.00),
        ("TOTAL REVENUE", 330000.00),
    ]

    for item, value in revenue_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Direct Costs
    ws[f'A{row}'] = "DIRECT COSTS"
    ws[f'A{row}'].font = section_font
    row += 1

    direct_costs = [
        ("Contractor Labor", 80000.00),
        ("Materials & Supplies", 20000.00),
        ("Travel Expenses", 15000.00),
        ("TOTAL DIRECT COSTS", 115000.00),
    ]

    for item, value in direct_costs:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C8-C12"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C14/C8"
    ws[f'E{row}'].number_format = percent_format
    row += 2

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = section_font
    row += 1

    op_expenses = [
        ("Salaries & Wages", 90000.00),
        ("Office Rent", 24000.00),
        ("Professional Services", 10000.00),
        ("Software Subscriptions", 8000.00),
        ("Marketing", 12000.00),
        ("Insurance", 6000.00),
        ("Utilities", 3000.00),
        ("TOTAL OPERATING EXPENSES", 153000.00),
    ]

    for item, value in op_expenses:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Operating Income
    ws[f'A{row}'] = "OPERATING INCOME"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C14-C22"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Other Items
    ws[f'A{row}'] = "OTHER INCOME/EXPENSE"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Interest Income"
    ws[f'C{row}'] = 2000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Interest Expense"
    ws[f'C{row}'] = -3000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C24+C25+C26"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C27/C8"
    ws[f'E{row}'].number_format = percent_format

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 15

    return wb

def create_restaurant_pandl_template():
    """Create Restaurant P&L template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Restaurant P&L"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "RESTAURANT PROFIT AND LOSS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Sales Section
    row = 3
    ws[f'A{row}'] = "SALES"
    ws[f'A{row}'].font = section_font
    row += 1

    sales_items = [
        ("Food Sales", 200000.00),
        ("Beverage Sales", 80000.00),
        ("TOTAL SALES", 280000.00),
    ]

    for item, value in sales_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # COGS Section
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = section_font
    row += 1

    cogs_items = [
        ("Food Cost", 60000.00),
        ("Beverage Cost", 16000.00),
        ("TOTAL COGS", 76000.00),
    ]

    for item, value in cogs_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C7-C10"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C11/C7"
    ws[f'E{row}'].number_format = percent_format
    row += 2

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = section_font
    row += 1

    op_expenses = [
        ("Labor Costs", 70000.00),
        ("Rent", 15000.00),
        ("Utilities", 5000.00),
        ("Marketing", 3000.00),
        ("Supplies", 4000.00),
        ("Repairs & Maintenance", 2000.00),
        ("Insurance", 3000.00),
        ("TOTAL OPERATING EXPENSES", 102000.00),
    ]

    for item, value in op_expenses:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Operating Income
    ws[f'A{row}'] = "OPERATING INCOME"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C11-C19"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C20"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format
    ws[f'E{row}'] = "=C21/C7"
    ws[f'E{row}'].number_format = percent_format

    # Key Metrics
    row += 2
    ws[f'A{row}'] = "KEY RESTAURANT METRICS"
    ws[f'A{row}'].font = section_font
    row += 1

    metrics = [
        ("Food Cost %", "=C8/C7"),
        ("Beverage Cost %", "=C9/C7"),
        ("Labor Cost %", "=C13/C7"),
        ("Prime Cost %", "=(C8+C9+C13)/C7"),
        ("Occupancy Cost %", "=(C14+C15)/C7"),
    ]

    for metric, formula in metrics:
        ws[f'A{row}'] = metric
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = percent_format
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 15

    return wb

def create_startup_financials_template():
    """Create Startup Financial Model template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Startup Financials"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    months = ["Item", "Month 1", "Month 2", "Month 3", "Q1 Total", 
              "Month 4", "Month 5", "Month 6", "Q2 Total",
              "Month 7", "Month 8", "Month 9", "Q3 Total",
              "Month 10", "Month 11", "Month 12", "Q4 Total", "Year Total"]

    for col, month in enumerate(months, start=1):
        cell = ws.cell(row=1, column=col, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Revenue Section
    row = 2
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    revenue_items = [
        ["Product Sales", 5000, 8000, 12000, 0, 15000, 18000, 22000, 0, 25000, 28000, 32000, 0, 35000, 38000, 42000, 0, 0],
        ["Service Revenue", 3000, 5000, 7000, 0, 9000, 11000, 13000, 0, 15000, 17000, 19000, 0, 21000, 23000, 25000, 0, 0],
    ]

    for row_data in revenue_items:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
        row += 1

    # Add formulas for quarterly and yearly totals
    for r in range(3, 5):
        for q_start, q_end, q_col in [(2, 4, 5), (6, 8, 9), (10, 12, 13), (14, 16, 17)]:
            col_letter = get_column_letter(q_col)
            ws[f'{col_letter}{r}'] = f'=SUM({get_column_letter(q_start)}{r}:{get_column_letter(q_end)}{r})'
        ws[f'R{r}'] = f'=SUM(B{r}:Q{r})'

    # COGS Section
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    cogs_items = [
        ["Direct Materials", 2000, 3200, 4800, 0, 6000, 7200, 8800, 0, 10000, 11200, 12800, 0, 14000, 15200, 16800, 0, 0],
        ["Direct Labor", 1000, 1600, 2400, 0, 3000, 3600, 4400, 0, 5000, 5600, 6400, 0, 7000, 7600, 8400, 0, 0],
    ]

    for row_data in cogs_items:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
        row += 1

    # Add formulas
    for r in range(6, 8):
        for q_start, q_end, q_col in [(2, 4, 5), (6, 8, 9), (10, 12, 13), (14, 16, 17)]:
            col_letter = get_column_letter(q_col)
            ws[f'{col_letter}{r}'] = f'=SUM({get_column_letter(q_start)}{r}:{get_column_letter(q_end)}{r})'
        ws[f'R{r}'] = f'=SUM(B{r}:Q{r})'

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 19):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{row}'] = f'=SUM({col_letter}3:{col_letter}5)'
    row += 1

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    expenses = [
        ["Salaries", 10000, 10000, 10000, 0, 12000, 12000, 12000, 0, 15000, 15000, 15000, 0, 18000, 18000, 18000, 0, 0],
        ["Rent", 3000, 3000, 3000, 0, 3000, 3000, 3000, 0, 3000, 3000, 3000, 0, 3000, 3000, 3000, 0, 0],
        ["Marketing", 2000, 3000, 4000, 0, 5000, 6000, 7000, 0, 8000, 9000, 10000, 0, 11000, 12000, 13000, 0, 0],
    ]

    for row_data in expenses:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
        row += 1

    # Add formulas
    for r in range(11, 14):
        for q_start, q_end, q_col in [(2, 4, 5), (6, 8, 9), (10, 12, 13), (14, 16, 17)]:
            col_letter = get_column_letter(q_col)
            ws[f'{col_letter}{r}'] = f'=SUM({get_column_letter(q_start)}{r}:{get_column_letter(q_end)}{r})'
        ws[f'R{r}'] = f'=SUM(B{r}:Q{r})'

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    for col in range(2, 19):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{row}'] = f'={col_letter}8-SUM({col_letter}10:{col_letter}13)'

    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 19):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Format as currency
    for r in range(2, 15):
        for c in range(2, 19):
            ws.cell(row=r, column=c).number_format = '#,##0'

    return wb

def create_real_estate_template():
    """Create Real Estate Investment Analysis template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Real Estate Analysis"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "REAL ESTATE INVESTMENT ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Property Information
    row = 3
    ws[f'A{row}'] = "PROPERTY INFORMATION"
    ws[f'A{row}'].font = section_font
    row += 1

    prop_info = [
        ("Purchase Price", 500000.00),
        ("Closing Costs", 15000.00),
        ("Renovation Costs", 25000.00),
        ("Total Investment", 540000.00),
    ]

    for item, value in prop_info:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "Total" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Financing
    row += 1
    ws[f'A{row}'] = "FINANCING"
    ws[f'A{row}'].font = section_font
    row += 1

    financing = [
        ("Down Payment (20%)", 100000.00),
        ("Loan Amount", 400000.00),
        ("Interest Rate", 0.065),
        ("Loan Term (Years)", 30),
        ("Monthly Payment", 2528.00),
    ]

    for item, value in financing:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        if "Rate" in item or "Term" in item:
            ws[f'C{row}'].number_format = percent_format if "Rate" in item else '0'
        else:
            ws[f'C{row}'].number_format = currency_format
        row += 1

    # Rental Income
    row += 1
    ws[f'A{row}'] = "RENTAL INCOME"
    ws[f'A{row}'].font = section_font
    row += 1

    rental = [
        ("Monthly Rent", 3500.00),
        ("Annual Gross Income", 42000.00),
        ("Vacancy Rate (5%)", 0.05),
        ("Vacancy Loss", 2100.00),
        ("Effective Gross Income", 39900.00),
    ]

    for item, value in rental:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        if "Rate" in item:
            ws[f'C{row}'].number_format = percent_format
        else:
            ws[f'C{row}'].number_format = currency_format
        if "Effective" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Operating Expenses
    row += 1
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = section_font
    row += 1

    op_expenses = [
        ("Property Tax", 6000.00),
        ("Insurance", 2400.00),
        ("Maintenance", 3000.00),
        ("Property Management", 2400.00),
        ("Utilities", 1800.00),
        ("Total Operating Expenses", 15600.00),
    ]

    for item, value in op_expenses:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "Total" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # NOI and Returns
    row += 1
    ws[f'A{row}'] = "INVESTMENT RETURNS"
    ws[f'A{row}'].font = section_font
    row += 1

    returns = [
        ("Net Operating Income (NOI)", 24300.00),
        ("Cap Rate", 0.045),
        ("Cash Flow Before Tax", 13964.00),
        ("Cash on Cash Return", 0.0259),
    ]

    for item, value in returns:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        if "Rate" in item or "Return" in item:
            ws[f'C{row}'].number_format = percent_format
        else:
            ws[f'C{row}'].number_format = currency_format
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    return wb

def main():
    """Generate all specialized financial templates"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"

    print("Generating Specialized Financial Excel Templates...")
    print("=" * 60)

    templates = [
        ("Retail_PandL.xlsx", create_retail_pandl_template),
        ("Service_Business_PandL.xlsx", create_service_business_pandl_template),
        ("Restaurant_PandL.xlsx", create_restaurant_pandl_template),
        ("Startup_Financials.xlsx", create_startup_financials_template),
        ("Real_Estate_Analysis.xlsx", create_real_estate_template),
    ]

    for filename, create_func in templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_func()
        wb.save(filepath)
        print(f"✓ Created: {filename}")

    print("=" * 60)
    print(f"\nAll templates saved to: {output_dir}")
    print("\nSpecialized templates include:")
    print("  • Retail P&L Statement")
    print("  • Service Business P&L")
    print("  • Restaurant P&L")
    print("  • Startup Financial Model")
    print("  • Real Estate Investment Analysis")

if __name__ == "__main__":
    main()
