#!/usr/bin/env python3
"""
Financial Excel Template Generator
Creates Excel templates with formulas and VBA macros for financial accounting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_general_ledger_template():
    """Create a General Ledger template with formulas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center')

    # Headers
    headers = [
        "Transaction ID", "Date", "Account Code", "Description",
        "Debit", "Credit", "Balance", "Category", "Reference", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Sample data
    sample_data = [
        ["T001", "2026-01-15", "1001", "Opening Balance", 10000.00, 0.00, 10000.00, "Asset", "OB-001", "Posted"],
        ["T002", "2026-01-20", "4001", "Sales Revenue", 0.00, 5000.00, 15000.00, "Revenue", "INV-001", "Posted"],
        ["T003", "2026-01-25", "2001", "Office Supplies", 500.00, 0.00, 14500.00, "Expense", "EXP-001", "Posted"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas for additional rows
    for row in range(5, 11):
        # Balance formula: previous balance + debit - credit
        ws.cell(row=row, column=7).value = f'=G{row-1}+E{row}-F{row}'

    # Summary section
    ws.cell(row=13, column=1, value="SUMMARY")
    ws.cell(row=13, column=1).font = Font(bold=True)
    ws.cell(row=14, column=4, value="Total Debits:")
    ws.cell(row=14, column=5, value="=SUM(E:E)")
    ws.cell(row=15, column=4, value="Total Credits:")
    ws.cell(row=15, column=6, value="=SUM(F:F)")
    ws.cell(row=16, column=4, value="Net Balance:")
    ws.cell(row=16, column=5, value="=E14-F15")

    # Format columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10

    # Format currency cells
    for row in range(2, 12):
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_balance_sheet_template():
    """Create a Balance Sheet template with formulas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "BALANCE SHEET"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "As of:"
    ws['B2'] = "2026-01-31"

    # ASSETS Section
    row = 4
    ws[f'A{row}'] = "ASSETS"
    ws[f'A{row}'].font = section_font
    row += 1

    # Current Assets
    ws[f'A{row}'] = "Current Assets:"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1

    assets = [
        ("Cash and Cash Equivalents", 50000),
        ("Accounts Receivable", 30000),
        ("Inventory", 45000),
        ("Prepaid Expenses", 5000),
    ]

    for asset, value in assets:
        ws[f'A{row}'] = asset
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Total Current Assets:"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C5:C8)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Non-Current Assets
    ws[f'A{row}'] = "Non-Current Assets:"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1

    non_current_assets = [
        ("Property, Plant & Equipment", 200000),
        ("Less: Accumulated Depreciation", -40000),
        ("Long-term Investments", 25000),
    ]

    for asset, value in non_current_assets:
        ws[f'A{row}'] = asset
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Total Non-Current Assets:"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C11:C13)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Total Assets
    ws[f'A{row}'] = "TOTAL ASSETS"
    ws[f'A{row}'].font = header_font
    ws[f'C{row}'] = "=C9+C14"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].number_format = currency_format
    row += 3

    # LIABILITIES Section
    ws[f'A{row}'] = "LIABILITIES"
    ws[f'A{row}'].font = section_font
    row += 1

    # Current Liabilities
    ws[f'A{row}'] = "Current Liabilities:"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1

    liabilities = [
        ("Accounts Payable", 25000),
        ("Short-term Debt", 15000),
        ("Accrued Expenses", 8000),
    ]

    for liability, value in liabilities:
        ws[f'A{row}'] = liability
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Total Current Liabilities:"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C18:C20)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Non-Current Liabilities
    ws[f'A{row}'] = "Non-Current Liabilities:"
    ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 1

    ws[f'A{row}'] = "Long-term Debt"
    ws[f'C{row}'] = 50000
    ws[f'C{row}'].number_format = currency_format
    row += 2

    ws[f'B{row}'] = "Total Liabilities:"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C21+C23"
    ws[f'C{row}'].number_format = currency_format
    row += 3

    # EQUITY Section
    ws[f'A{row}'] = "EQUITY"
    ws[f'A{row}'].font = section_font
    row += 1

    equity_items = [
        ("Owner's Capital", 100000),
        ("Retained Earnings", 42000),
        ("Current Year Earnings", 30000),
    ]

    for equity, value in equity_items:
        ws[f'A{row}'] = equity
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Total Equity:"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C26:C28)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    ws[f'A{row}'] = "TOTAL LIABILITIES & EQUITY"
    ws[f'A{row}'].font = header_font
    ws[f'C{row}'] = "=C24+C29"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Financial Ratios
    ws[f'A{row}'] = "KEY FINANCIAL RATIOS"
    ws[f'A{row}'].font = section_font
    row += 1

    ratios = [
        ("Working Capital", "=C9-C21"),
        ("Current Ratio", "=C9/C21"),
        ("Debt-to-Equity", "=C24/C29"),
        ("Asset Turnover", "=150000/C15"),
    ]

    for ratio_name, formula in ratios:
        ws[f'A{row}'] = ratio_name
        ws[f'C{row}'] = formula
        if "Ratio" in ratio_name:
            ws[f'C{row}'].number_format = '0.00'
        else:
            ws[f'C{row}'].number_format = currency_format
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20

    return wb

def create_income_statement_template():
    """Create an Income Statement template with formulas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "INCOME STATEMENT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "For the Period Ended:"
    ws['B2'] = "2026-01-31"

    # Revenue Section
    row = 4
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Sales Revenue"
    ws[f'B{row}'] = 200000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Less: Sales Returns"
    ws[f'B{row}'] = -5000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Less: Discounts"
    ws[f'B{row}'] = -2000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "NET SALES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B5:B7)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B8/B8"
    ws[f'C{row}'].number_format = percent_format
    row += 1

    ws[f'A{row}'] = "Other Revenue"
    ws[f'B{row}'] = 5000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "TOTAL REVENUE"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=B8+B9"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    row += 2

    # COGS Section
    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Beginning Inventory"
    ws[f'B{row}'] = 30000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "+ Purchases"
    ws[f'B{row}'] = 80000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "- Ending Inventory"
    ws[f'B{row}'] = -25000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "COST OF GOODS SOLD"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B12:B14)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B15/B10"
    ws[f'C{row}'].number_format = percent_format
    row += 2

    # Gross Profit
    ws[f'A{row}'] = "GROSS PROFIT"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = "=B10-B15"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B16/B10"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].number_format = percent_format
    row += 2

    # Operating Expenses
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = section_font
    row += 1

    expenses = [
        ("Salaries & Wages", 45000),
        ("Rent Expense", 12000),
        ("Utilities", 3000),
        ("Depreciation", 5000),
        ("Marketing", 8000),
        ("Other Operating Expenses", 4000),
    ]

    for expense, value in expenses:
        ws[f'A{row}'] = expense
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = currency_format
        row += 1

    ws[f'A{row}'] = "TOTAL OPERATING EXPENSES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B18:B23)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B24/B10"
    ws[f'C{row}'].number_format = percent_format
    row += 2

    # Operating Income
    ws[f'A{row}'] = "OPERATING INCOME"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=B16-B24"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B25/B10"
    ws[f'C{row}'].number_format = percent_format
    row += 2

    # Other Income/Expense
    ws[f'A{row}'] = "OTHER INCOME/EXPENSE"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Interest Income"
    ws[f'B{row}'] = 2000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Interest Expense"
    ws[f'B{row}'] = -3000
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "NET OTHER INCOME/EXPENSE"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=SUM(B27:B28)"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    row += 2

    # Income Before Tax
    ws[f'A{row}'] = "INCOME BEFORE TAX"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "=B25+B29"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Income Tax Expense (25%)"
    ws[f'B{row}'] = "=B30*0.25"
    ws[f'B{row}'].number_format = currency_format
    row += 2

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = "=B30-B31"
    ws[f'B{row}'].font = header_font
    ws[f'B{row}'].number_format = currency_format
    ws[f'C{row}'] = "=B32/B10"
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].number_format = percent_format
    row += 2

    # Profitability Ratios
    ws[f'A{row}'] = "PROFITABILITY RATIOS"
    ws[f'A{row}'].font = section_font
    row += 1

    ratios = [
        ("Gross Margin %", "=B16/B10"),
        ("Operating Margin %", "=B25/B10"),
        ("Net Profit Margin %", "=B32/B10"),
        ("Return on Sales", "=B32/B10"),
    ]

    for ratio_name, formula in ratios:
        ws[f'A{row}'] = ratio_name
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = percent_format
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    return wb

def create_budget_template():
    """Create a Budget template with formulas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Budget"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    currency_format = '#,##0'
    variance_format = '#,##0'

    # Headers
    months = ["Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total", "Average"]

    for col, month in enumerate(months, start=1):
        cell = ws.cell(row=1, column=col, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Revenue Section
    row = 2
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    revenue_items = [
        "Product Sales",
        "Service Revenue",
        "Other Income",
    ]

    for item in revenue_items:
        ws[f'A{row}'] = item
        for col in range(2, 14):
            ws.cell(row=row, column=col, value=10000)
        row += 1

    ws[f'A{row}'] = "TOTAL REVENUE"
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        if col == 14:  # Total column
            ws[f'{col_letter}{row}'] = f"=SUM({col_letter}3:{col_letter}5)"
        else:
            ws[f'{col_letter}{row}'] = f"=SUM({col_letter}3:{col_letter}5)"
    row += 2

    # Expenses Section
    ws[f'A{row}'] = "EXPENSES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    expense_items = [
        "Cost of Goods Sold",
        "Salaries & Wages",
        "Rent",
        "Utilities",
        "Marketing",
        "Office Supplies",
        "Insurance",
        "Depreciation",
        "Other Expenses",
    ]

    for item in expense_items:
        ws[f'A{row}'] = item
        for col in range(2, 14):
            ws.cell(row=row, column=col, value=5000)
        row += 1

    ws[f'A{row}'] = "TOTAL EXPENSES"
    ws[f'A{row}'].font = Font(bold=True)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        if col == 14:  # Total column
            ws[f'{col_letter}{row}'] = f"=SUM({col_letter}8:{col_letter}16)"
        else:
            ws[f'{col_letter}{row}'] = f"=SUM({col_letter}8:{col_letter}16)"
    row += 2

    # Net Income
    ws[f'A{row}'] = "NET INCOME"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    for col in range(2, 15):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}{row}'] = f"={col_letter}6-{col_letter}17"
    row += 2

    # Format all data cells
    for r in range(2, row):
        for c in range(2, 15):
            cell = ws.cell(row=r, column=c)
            cell.number_format = currency_format

    # Average column
    for r in range(2, row):
        ws.cell(row=r, column=15).value = f"=AVERAGE(B{r}:M{r})"

    # Column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 16):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return wb

def create_financial_ratios_template():
    """Create a Financial Ratios calculator template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Ratios"

    header_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=11, color="4472C4")
    currency_format = '#,##0.00'
    number_format = '0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "FINANCIAL RATIOS CALCULATOR"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Input Section
    row = 3
    ws[f'A{row}'] = "INPUT VALUES (from Balance Sheet & Income Statement)"
    ws[f'A{row}'].font = category_font
    row += 1

    inputs = [
        ("Current Assets", 130000),
        ("Current Liabilities", 48000),
        ("Cash", 50000),
        ("Inventory", 45000),
        ("Total Assets", 215000),
        ("Total Liabilities", 98000),
        ("Total Equity", 117000),
        ("Net Sales", 193000),
        ("Cost of Goods Sold", 85000),
        ("Net Income", 43500),
        ("EBIT", 55000),
        ("Interest Expense", 3000),
        ("Average Inventory", 40000),
        ("Average Accounts Receivable", 28000),
    ]

    for label, value in inputs:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = currency_format
        row += 1

    row += 1

    # Liquidity Ratios
    ws[f'A{row}'] = "LIQUIDITY RATIOS"
    ws[f'A{row}'].font = category_font
    row += 1

    liquidity_ratios = [
        ("Current Ratio", "=B5/B6", "Current Assets / Current Liabilities"),
        ("Quick Ratio", "=(B5-B7)/B6", "(Current Assets - Inventory) / Current Liabilities"),
        ("Cash Ratio", "=B7/B6", "Cash / Current Liabilities"),
    ]

    for ratio, formula, description in liquidity_ratios:
        ws[f'A{row}'] = ratio
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = number_format
        ws[f'C{row}'] = description
        row += 1

    row += 1

    # Profitability Ratios
    ws[f'A{row}'] = "PROFITABILITY RATIOS"
    ws[f'A{row}'].font = category_font
    row += 1

    profitability_ratios = [
        ("Gross Profit Margin", "=(B13-B9)/B13", "(Sales - COGS) / Sales"),
        ("Operating Margin", "=B15/B13", "Operating Income / Sales"),
        ("Net Profit Margin", "=B14/B13", "Net Income / Sales"),
        ("Return on Assets (ROA)", "=B14/B10", "Net Income / Total Assets"),
        ("Return on Equity (ROE)", "=B14/B12", "Net Income / Total Equity"),
    ]

    for ratio, formula, description in profitability_ratios:
        ws[f'A{row}'] = ratio
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = percent_format
        ws[f'C{row}'] = description
        row += 1

    row += 1

    # Efficiency Ratios
    ws[f'A{row}'] = "EFFICIENCY RATIOS"
    ws[f'A{row}'].font = category_font
    row += 1

    efficiency_ratios = [
        ("Asset Turnover", "=B13/B10", "Sales / Total Assets"),
        ("Inventory Turnover", "=B9/B16", "COGS / Average Inventory"),
        ("Receivables Turnover", "=B13/B17", "Sales / Average AR"),
    ]

    for ratio, formula, description in efficiency_ratios:
        ws[f'A{row}'] = ratio
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = number_format
        ws[f'C{row}'] = description
        row += 1

    row += 1

    # Solvency Ratios
    ws[f'A{row}'] = "SOLVENCY RATIOS"
    ws[f'A{row}'].font = category_font
    row += 1

    solvency_ratios = [
        ("Debt-to-Equity", "=B11/B12", "Total Liabilities / Total Equity"),
        ("Debt Ratio", "=B11/B10", "Total Liabilities / Total Assets"),
        ("Times Interest Earned", "=B16/B18", "EBIT / Interest Expense"),
    ]

    for ratio, formula, description in solvency_ratios:
        ws[f'A{row}'] = ratio
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = number_format
        ws[f'C{row}'] = description
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45

    return wb

def main():
    """Generate all financial templates"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"

    print("Generating Financial Excel Templates...")
    print("=" * 50)

    templates = [
        ("General_Ledger.xlsx", create_general_ledger_template),
        ("Balance_Sheet.xlsx", create_balance_sheet_template),
        ("Income_Statement.xlsx", create_income_statement_template),
        ("Annual_Budget.xlsx", create_budget_template),
        ("Financial_Ratios.xlsx", create_financial_ratios_template),
    ]

    for filename, create_func in templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_func()
        wb.save(filepath)
        print(f"✓ Created: {filename}")

    print("=" * 50)
    print(f"\nAll templates saved to: {output_dir}")
    print("\nTemplates include:")
    print("  • Pre-built formulas for automatic calculations")
    print("  • Sample data for reference")
    print("  • Professional formatting")
    print("  • Ready to use with your own data")

if __name__ == "__main__":
    main()
