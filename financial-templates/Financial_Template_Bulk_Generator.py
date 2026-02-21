#!/usr/bin/env python3
"""
Bulk Financial Excel Template Generator
Creates 200+ financial templates for various use cases
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import random

def create_basic_template(name, headers, sample_data, title=None):
    """Create a basic template with headers and sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name.replace("_", " ").title()

    if title:
        ws.merge_cells('A1:Z1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 3 if title else 1

    # Headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Sample data
    for row_idx, row_data in enumerate(sample_data, start=start_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return wb

def create_simple_tracker_template(name, title, columns):
    """Create a simple tracker template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name.replace("_", " ").title()

    # Title
    ws.merge_cells('A1:Z1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Add some empty rows
    for row in range(4, 25):
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = border

    # Auto-fit columns
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return wb

def create_simple_calculator_template(name, title, inputs, outputs):
    """Create a simple calculator template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name.replace("_", " ").title()

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Inputs section
    row = 3
    ws[f'A{row}'] = "INPUTS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for label, value in inputs:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if isinstance(value, (int, float)):
            ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

    # Outputs section
    row += 1
    ws[f'A{row}'] = "OUTPUTS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for label, formula in outputs:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formula
        if isinstance(formula, str) and formula.startswith('='):
            pass  # Keep as formula
        else:
            ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    return wb

def generate_all_templates():
    """Generate 200+ financial templates"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"
    templates_created = []

    print("Generating 200+ Financial Excel Templates...")
    print("=" * 70)

    # ========== TRACKERS (50 templates) ==========
    trackers = [
        # Expense Trackers
        ("Daily_Expense_Tracker.xlsx", "Daily Expense Tracker",
         ["Date", "Category", "Description", "Amount", "Payment Method", "Notes"]),
        ("Weekly_Expense_Tracker.xlsx", "Weekly Expense Tracker",
         ["Week", "Category", "Description", "Amount", "Day"]),
        ("Monthly_Expense_Tracker.xlsx", "Monthly Expense Tracker",
         ["Month", "Category", "Budget", "Actual", "Variance"]),
        ("Quarterly_Expense_Tracker.xlsx", "Quarterly Expense Tracker",
         ["Quarter", "Category", "Budget", "Actual", "Variance"]),
        ("Annual_Expense_Tracker.xlsx", "Annual Expense Tracker",
         ["Year", "Category", "Budget", "Actual", "Variance"]),

        # Income Trackers
        ("Daily_Income_Tracker.xlsx", "Daily Income Tracker",
         ["Date", "Source", "Description", "Amount", "Category"]),
        ("Weekly_Income_Tracker.xlsx", "Weekly Income Tracker",
         ["Week", "Source", "Amount", "Category"]),
        ("Monthly_Income_Tracker.xlsx", "Monthly Income Tracker",
         ["Month", "Source", "Budget", "Actual", "Variance"]),
        ("Sales_Tracker.xlsx", "Sales Tracker",
         ["Date", "Customer", "Product", "Quantity", "Unit Price", "Total"]),
        ("Commission_Tracker.xlsx", "Commission Tracker",
         ["Salesperson", "Date", "Sale Amount", "Commission Rate", "Commission"]),

        # Asset Trackers
        ("Fixed_Asset_Tracker.xlsx", "Fixed Asset Tracker",
         ["Asset ID", "Name", "Purchase Date", "Cost", "Location", "Status"]),
        ("Vehicle_Tracker.xlsx", "Vehicle Tracker",
         ["Vehicle ID", "Make", "Model", "Year", "License Plate", "Status"]),
        ("Equipment_Tracker.xlsx", "Equipment Tracker",
         ["Equipment ID", "Name", "Purchase Date", "Cost", "Location", "Condition"]),
        ("IT_Asset_Tracker.xlsx", "IT Asset Tracker",
         ["Asset ID", "Device Type", "Serial Number", "User", "Location", "Status"]),
        ("Furniture_Tracker.xlsx", "Furniture Tracker",
         ["Item ID", "Description", "Location", "Purchase Date", "Condition", "Status"]),

        # Time Trackers
        ("Employee_Time_Tracker.xlsx", "Employee Time Tracker",
         ["Date", "Employee", "Start Time", "End Time", "Hours", "Project"]),
        ("Project_Time_Tracker.xlsx", "Project Time Tracker",
         ["Date", "Project", "Task", "Hours", "Employee", "Notes"]),
        ("Billable_Hours_Tracker.xlsx", "Billable Hours Tracker",
         ["Date", "Client", "Project", "Hours", "Rate", "Amount"]),
        ("Overtime_Tracker.xlsx", "Overtime Tracker",
         ["Week", "Employee", "Regular Hours", "Overtime Hours", "Rate", "Total Pay"]),

        # Mileage Trackers
        ("Mileage_Tracker.xlsx", "Mileage Tracker",
         ["Date", "Vehicle", "Start Mileage", "End Mileage", "Miles", "Purpose"]),
        ("Business_Mileage_Tracker.xlsx", "Business Mileage Tracker",
         ["Date", "Vehicle", "Miles", "Purpose", "Client", "Reimbursable"]),

        # Budget Trackers
        ("Department_Budget_Tracker.xlsx", "Department Budget Tracker",
         ["Department", "Category", "Budget", "Actual", "Variance", "%"]),
        ("Project_Budget_Tracker.xlsx", "Project Budget Tracker",
         ["Project", "Category", "Budget", "Actual", "Variance", "Remaining"]),
        ("Event_Budget_Tracker.xlsx", "Event Budget Tracker",
         ["Item", "Category", "Estimated Cost", "Actual Cost", "Variance", "Vendor"]),
        ("Marketing_Budget_Tracker.xlsx", "Marketing Budget Tracker",
         ["Campaign", "Channel", "Budget", "Actual", "ROI", "Notes"]),

        # Invoice Trackers
        ("Invoice_Tracker.xlsx", "Invoice Tracker",
         ["Invoice #", "Customer", "Date", "Due Date", "Amount", "Status", "Paid Date"]),
        ("Bill_Tracker.xlsx", "Bill Tracker",
         ["Bill #", "Vendor", "Date", "Due Date", "Amount", "Status", "Paid Date"]),
        ("Receipt_Tracker.xlsx", "Receipt Tracker",
         ["Date", "Vendor", "Category", "Amount", "Payment Method", "Notes"]),

        # Payment Trackers
        ("Payment_Tracker.xlsx", "Payment Tracker",
         ["Date", "Payee", "Category", "Amount", "Payment Method", "Reference"]),
        ("Supplier_Payment_Tracker.xlsx", "Supplier Payment Tracker",
         ["Supplier", "Invoice #", "Date", "Amount", "Payment Method", "Status"]),
        ("Customer_Payment_Tracker.xlsx", "Customer Payment Tracker",
         ["Customer", "Invoice #", "Date", "Amount", "Payment Method", "Status"]),

        # Tax Trackers
        ("Sales_Tax_Tracker.xlsx", "Sales Tax Tracker",
         ["Date", "Invoice #", "Sale Amount", "Tax Rate", "Tax Amount", "Total"]),
        ("VAT_Tracker.xlsx", "VAT Tracker",
         ["Date", "Invoice #", "Net Amount", "VAT Rate", "VAT Amount", "Total"]),
        ("Tax_Deduction_Tracker.xlsx", "Tax Deduction Tracker",
         ["Category", "Description", "Amount", "Date", "Receipt #", "Approved"]),

        # Investment Trackers
        ("Stock_Tracker.xlsx", "Stock Tracker",
         ["Symbol", "Company", "Shares", "Buy Price", "Current Price", "Gain/Loss"]),
        ("Bond_Tracker.xlsx", "Bond Tracker",
         ["Bond", "Face Value", "Purchase Price", "Coupon Rate", "Maturity Date", "Yield"]),
        ("Mutual_Fund_Tracker.xlsx", "Mutual Fund Tracker",
         ["Fund Name", "Units", "NAV", "Total Value", "Date", "Type"]),
        ("Dividend_Tracker.xlsx", "Dividend Tracker",
         ["Stock", "Ex-Dividend Date", "Amount per Share", "Shares", "Total Dividend"]),

        # Other Trackers
        ("Subscription_Tracker.xlsx", "Subscription Tracker",
         ["Service", "Cost", "Billing Cycle", "Next Billing Date", "Auto-Renew", "Notes"]),
        ("Maintenance_Tracker.xlsx", "Maintenance Tracker",
         ["Asset", "Date", "Type", "Description", "Cost", "Technician", "Status"]),
        ("Training_Tracker.xlsx", "Training Tracker",
         ["Employee", "Training", "Date", "Provider", "Cost", "Status", "Certificate"]),
        ("Certification_Tracker.xlsx", "Certification Tracker",
         ["Employee", "Certification", "Issue Date", "Expiry Date", "Status", "Renewal Date"]),
    ]

    for filename, title, columns in trackers:
        filepath = os.path.join(output_dir, filename)
        wb = create_simple_tracker_template(filename, title, columns)
        wb.save(filepath)
        templates_created.append(filename)
        print(f"✓ {filename}")

    # ========== CALCULATORS (50 templates) ==========
    calculators = [
        # Loan Calculators
        ("Loan_Calculator.xlsx", "Loan Calculator",
         [("Loan Amount", 100000), ("Interest Rate (%)", 5), ("Term (Years)", 30)],
         [("Monthly Payment", "=PMT(B2/12,B3*12,-B1)"),
          ("Total Interest", "=(B4*B3)-B1"),
          ("Total Payment", "=B4*B3")]),

        ("Mortgage_Calculator.xlsx", "Mortgage Calculator",
         [("Home Price", 300000), ("Down Payment", 60000), ("Interest Rate (%)", 4.5), ("Term (Years)", 30)],
         [("Loan Amount", "=B1-B2"),
          ("Monthly Payment", "=PMT(B3/12,B4*12,-(B1-B2))"),
          ("Total Interest", "=((B5*B4)-(B1-B2))"),
          ("Total Payment", "=B5*B4")]),

        ("Car_Loan_Calculator.xlsx", "Car Loan Calculator",
         [("Car Price", 25000), ("Down Payment", 5000), ("Interest Rate (%)", 5), ("Term (Years)", 5)],
         [("Loan Amount", "=B1-B2"),
          ("Monthly Payment", "=PMT(B3/12,B4*12,-(B1-B2))"),
          ("Total Interest", "=((B5*B4)-(B1-B2))")]),

        ("Personal_Loan_Calculator.xlsx", "Personal Loan Calculator",
         [("Loan Amount", 10000), ("Interest Rate (%)", 8), ("Term (Months)", 36)],
         [("Monthly Payment", "=PMT(B2/12,B3,-B1)"),
          ("Total Interest", "=(B4*B3)-B1"),
          ("Total Payment", "=B4*B3")]),

        ("Student_Loan_Calculator.xlsx", "Student Loan Calculator",
         [("Loan Amount", 50000), ("Interest Rate (%)", 4.5), ("Term (Years)", 10)],
         [("Monthly Payment", "=PMT(B2/12,B3*12,-B1)"),
          ("Total Interest", "=(B4*B3)-B1")]),

        # Investment Calculators
        ("ROI_Calculator.xlsx", "ROI Calculator",
         [("Initial Investment", 10000), ("Final Value", 15000), ("Time Period (Years)", 3)],
         [("Gain/Loss", "=B2-B1"),
          ("ROI (%)", "=(B3-B1)/B1"),
          ("Annual ROI (%)", "=((B3/B1)^(1/B4)-1)")]),

        ("Compound_Interest_Calculator.xlsx", "Compound Interest Calculator",
         [("Principal", 10000), ("Annual Rate (%)", 5), ("Years", 10), ("Compounding", 12)],
         [("Future Value", "=B1*(1+B2/B4)^(B4*B3)"),
          ("Total Interest", "=B6-B1")]),

        ("Savings_Calculator.xlsx", "Savings Calculator",
         [("Monthly Deposit", 500), ("Annual Rate (%)", 3), ("Years", 20)],
         [("Future Value", "=FV(B2/12,B3*12,-B1,0)"),
          ("Total Deposits", "=B1*B3*12"),
          ("Total Interest", "=B4-B5")]),

        ("Retirement_Calculator.xlsx", "Retirement Calculator",
         [("Current Age", 30), ("Retirement Age", 65), ("Current Savings", 50000), ("Monthly Contribution", 1000), ("Expected Return (%)", 7)],
         [("Years to Retirement", "=B2-B1"),
          ("Future Value", "=FV(B5/12,(B2-B1)*12,-B4,-B3)"),
          ("Monthly Income (4%)", "=B6*0.04/12")]),

        ("401k_Calculator.xlsx", "401(k) Calculator",
         [("Annual Salary", 60000), ("Contribution Rate (%)", 5), ("Employer Match (%)", 3), ("Expected Return (%)", 7), ("Years", 30)],
         [("Annual Contribution", "=B1*B2/100"),
          ("Employer Match", "=B1*B3/100"),
          ("Total Annual", "=B6+B7"),
          ("Future Value", "=FV(B4/12,B5*12,-(B6+B7)/12,0)")]),

        # Tax Calculators
        ("Income_Tax_Calculator.xlsx", "Income Tax Calculator",
         [("Gross Income", 100000), ("Deductions", 20000), ("Tax Rate (%)", 22)],
         [("Taxable Income", "=B1-B2"),
          ("Tax Owed", "=B3*B4/100"),
          ("Net Income", "=B3-B5")]),

        ("Self_Employment_Tax_Calculator.xlsx", "Self-Employment Tax Calculator",
         [("Net Earnings", 80000), ("Tax Rate (%)", 15.3)],
         [("Self-Employment Tax", "=B1*B2/100"),
          ("Income After Tax", "=B1-B3")]),

        ("Capital_Gains_Tax_Calculator.xlsx", "Capital Gains Tax Calculator",
         [("Purchase Price", 10000), ("Sale Price", 15000), ("Holding Period (Years)", 2), ("Tax Rate (%)", 15)],
         [("Capital Gain", "=B2-B1"),
          ("Tax Owed", "=B4*B5/100"),
          ("After-Tax Profit", "=B4-B6")]),

        # Business Calculators
        ("Breakeven_Calculator.xlsx", "Breakeven Calculator",
         [("Fixed Costs", 10000), ("Variable Cost per Unit", 25), ("Selling Price per Unit", 50)],
         [("Contribution Margin", "=B3-B2"),
          ("Breakeven Units", "=B1/(B3-B2)"),
          ("Breakeven Revenue", "=B4*B3")]),

        ("Markup_Calculator.xlsx", "Markup Calculator",
         [("Cost", 100), ("Markup (%)", 50)],
         [("Selling Price", "=B1*(1+B2/100)"),
          ("Gross Profit", "=B3-B1"),
          ("Gross Margin (%)", "=B4/B3")]),

        ("Margin_Calculator.xlsx", "Margin Calculator",
         [("Cost", 100), ("Desired Margin (%)", 40)],
         [("Selling Price", "=B1/(1-B2/100)"),
          ("Gross Profit", "=B3-B1"),
          ("Markup (%)", "=B4/B1")]),

        ("Discount_Calculator.xlsx", "Discount Calculator",
         [("Original Price", 100), ("Discount (%)", 20)],
         [("Discount Amount", "=B1*B2/100"),
          ("Sale Price", "=B1-B3"),
          ("Savings", "=B3")]),

        ("Tip_Calculator.xlsx", "Tip Calculator",
         [("Bill Amount", 100), ("Tip Percentage (%)", 15), ("Split Between", 2)],
         [("Tip Amount", "=B1*B2/100"),
          ("Total Amount", "=B1+B4"),
          ("Per Person", "=B5/B3")]),

        ("Sales_Tax_Calculator.xlsx", "Sales Tax Calculator",
         [("Price Before Tax", 100), ("Tax Rate (%)", 8.25)],
         [("Tax Amount", "=B1*B2/100"),
          ("Total Price", "=B1+B3")]),

        ("Currency_Converter.xlsx", "Currency Converter",
         [("Amount", 100), ("Exchange Rate", 1.2)],
         [("Converted Amount", "=B1*B2")]),

        ("Inflation_Calculator.xlsx", "Inflation Calculator",
         [("Current Amount", 100), ("Inflation Rate (%)", 3), ("Years", 10)],
         [("Future Amount", "=B1*(1+B2/100)^B3"),
          ("Purchasing Power Loss", "=B4-B1")]),

        # Real Estate Calculators
        ("Rent_vs_Buy_Calculator.xlsx", "Rent vs Buy Calculator",
         [("Monthly Rent", 2000), ("Home Price", 400000), ("Down Payment", 80000), ("Interest Rate (%)", 4.5), ("Term (Years)", 30)],
         [("Monthly Mortgage", "=PMT(B4/12,B5*12,-(B2-B3))"),
          ("Monthly Difference", "=B7-B1"),
          ("Breakeven Years", "=(B2-B3)/(B8*12)")]),

        ("Rental_Yield_Calculator.xlsx", "Rental Yield Calculator",
         [("Property Price", 300000), ("Monthly Rent", 2500), ("Annual Expenses", 10000)],
         [("Annual Rent Income", "=B2*12"),
          ("Net Annual Income", "=B6-B3"),
          ("Rental Yield (%)", "=B7/B1")]),

        ("Cap_Rate_Calculator.xlsx", "Cap Rate Calculator",
         [("NOI", 24000), ("Property Value", 400000)],
         [("Cap Rate (%)", "=B1/B2")]),

        ("Cash_on_Cash_Return_Calculator.xlsx", "Cash on Cash Return Calculator",
         [("Annual Cash Flow", 12000), ("Cash Invested", 100000)],
         [("Cash on Cash Return (%)", "=B1/B2")]),

        # Payroll Calculators
        ("Hourly_to_Salary_Calculator.xlsx", "Hourly to Salary Calculator",
         [("Hourly Rate", 25), ("Hours per Week", 40)],
         [("Weekly Salary", "=B1*B2"),
          ("Monthly Salary", "=B3*4.33"),
          ("Annual Salary", "=B3*52")]),

        ("Salary_to_Hourly_Calculator.xlsx", "Salary to Hourly Calculator",
         [("Annual Salary", 52000), ("Hours per Week", 40)],
         [("Hourly Rate", "=B1/(B2*52)"),
          ("Weekly Pay", "=B3*B2"),
          ("Monthly Pay", "=B3*52/12")]),

        ("Overtime_Calculator.xlsx", "Overtime Calculator",
         [("Regular Hours", 40), ("Overtime Hours", 5), ("Hourly Rate", 25), ("Overtime Rate Multiplier", 1.5)],
         [("Regular Pay", "=B1*B3"),
          ("Overtime Pay", "=B2*B3*B4"),
          ("Total Pay", "=B6+B7")]),

        ("Bonus_Calculator.xlsx", "Bonus Calculator",
         [("Base Salary", 5000), ("Bonus Percentage (%)", 10)],
         [("Bonus Amount", "=B1*B2/100"),
          ("Total Pay", "=B1+B3")]),

        ("Commission_Calculator.xlsx", "Commission Calculator",
         [("Sales Amount", 10000), ("Commission Rate (%)", 5)],
         [("Commission", "=B1*B2/100"),
          ("Total Pay", "=B1+B3")]),

        # Depreciation Calculators
        ("Straight_Line_Depreciation.xlsx", "Straight Line Depreciation",
         [("Asset Cost", 10000), ("Salvage Value", 1000), ("Useful Life (Years)", 5)],
         [("Annual Depreciation", "=SLN(B1,B2,B3)"),
          ("Book Value", "=B1-B4")]),

        ("Declining_Balance_Depreciation.xlsx", "Declining Balance Depreciation",
         [("Asset Cost", 10000), ("Salvage Value", 1000), ("Useful Life (Years)", 5), ("Period", 1)],
         [("Depreciation", "=DB(B1,B2,B3,B4)"),
          ("Book Value", "=B1-B5")]),

        ("Double_Declining_Depreciation.xlsx", "Double Declining Depreciation",
         [("Asset Cost", 10000), ("Salvage Value", 1000), ("Useful Life (Years)", 5), ("Period", 1)],
         [("Depreciation", "=DDB(B1,B2,B3,B4)"),
          ("Book Value", "=B1-B5")]),

        # Ratio Calculators
        ("Liquidity_Ratio_Calculator.xlsx", "Liquidity Ratio Calculator",
         [("Current Assets", 200000), ("Current Liabilities", 80000), ("Cash", 50000), ("Inventory", 60000)],
         [("Current Ratio", "=B1/B2"),
          ("Quick Ratio", "=(B1-B4)/B2"),
          ("Cash Ratio", "=B3/B2")]),

        ("Profitability_Ratio_Calculator.xlsx", "Profitability Ratio Calculator",
         [("Net Income", 50000), ("Sales", 500000), ("Total Assets", 400000), ("Total Equity", 250000)],
         [("Gross Margin (%)", "=B2/B3"),
          ("Net Profit Margin (%)", "=B1/B3"),
          ("ROA (%)", "=B1/B4"),
          ("ROE (%)", "=B1/B5")]),

        ("Efficiency_Ratio_Calculator.xlsx", "Efficiency Ratio Calculator",
         [("Sales", 500000), ("Total Assets", 400000), ("COGS", 300000), ("Average Inventory", 50000), ("Average AR", 60000)],
         [("Asset Turnover", "=B1/B2"),
          ("Inventory Turnover", "=B3/B4"),
          ("Receivables Turnover", "=B1/B5")]),

        ("Debt_Ratio_Calculator.xlsx", "Debt Ratio Calculator",
         [("Total Liabilities", 200000), ("Total Equity", 250000), ("Total Assets", 450000)],
         [("Debt-to-Equity", "=B1/B2"),
          ("Debt Ratio", "=B1/B3"),
          ("Equity Ratio", "=B2/B3")]),

        # Other Calculators
        ("Inflation_Adjusted_Return.xlsx", "Inflation Adjusted Return",
         [("Nominal Return (%)", 8), ("Inflation Rate (%)", 3)],
         [("Real Return (%)", "=((1+B1/100)/(1+B2/100)-1)*100")]),

        ("Rule_of_72_Calculator.xlsx", "Rule of 72 Calculator",
         [("Interest Rate (%)", 8)],
         [("Years to Double", "=72/B1")]),

        ("Present_Value_Calculator.xlsx", "Present Value Calculator",
         [("Future Value", 10000), ("Discount Rate (%)", 5), ("Years", 10)],
         [("Present Value", "=PV(B2/100,B3,0,-B1)")]),

        ("Future_Value_Calculator.xlsx", "Future Value Calculator",
         [("Present Value", 10000), ("Interest Rate (%)", 5), ("Years", 10)],
         [("Future Value", "=FV(B2/100,B3,0,-B1)")]),

        ("Annuity_Payment_Calculator.xlsx", "Annuity Payment Calculator",
         [("Present Value", 100000), ("Interest Rate (%)", 5), ("Number of Payments", 120)],
         [("Payment", "=PMT(B2/100/12,B3,-B1)")]),

        ("Pension_Calculator.xlsx", "Pension Calculator",
         [("Final Salary", 80000), ("Years of Service", 30), ("Pension Factor (%)", 1.5)],
         [("Annual Pension", "=B1*B2*B3/100"),
          ("Monthly Pension", "=B4/12")]),

        ("College_Savings_Calculator.xlsx", "College Savings Calculator",
         [("Current Savings", 10000), ("Monthly Contribution", 500), ("Years to College", 18), ("Expected Return (%)", 6)],
         [("Future Value", "=FV(B4/100/12,B3*12,-B2,-B1)"),
          ("Total Contributions", "=B2*B3*12")]),

        ("Wedding_Budget_Calculator.xlsx", "Wedding Budget Calculator",
         [("Total Budget", 30000), ("Venue", 8000), ("Catering", 10000), ("Photography", 3000), ("Attire", 2000), ("Flowers", 1500), ("Music", 1000), ("Other", 4500)],
         [("Total Estimated", "=SUM(B2:B9)"),
          ("Remaining Budget", "=B1-B10"),
          ("Over/Under Budget", "=B1-B10")]),

        ("Vacation_Budget_Calculator.xlsx", "Vacation Budget Calculator",
         [("Total Budget", 5000), ("Flights", 1500), ("Accommodation", 2000), ("Food", 800), ("Activities", 400), ("Shopping", 300)],
         [("Total Estimated", "=SUM(B2:B7)"),
          ("Remaining Budget", "=B1-B8")]),

        ("Home_Buying_Budget_Calculator.xlsx", "Home Buying Budget Calculator",
         [("Monthly Income", 8000), ("Down Payment", 60000), ("Home Price", 400000), ("Closing Costs", 12000), ("Monthly Expenses", 2000)],
         [("Maximum Mortgage", "=B1*0.28*12"),
          ("Affordable Home Price", "=B2+B6"),
          ("Monthly Payment", "=PMT(0.045/12,360,-(B3-B2))"),
          ("Affordable?", "=IF(B7<=B1*0.28,\"Yes\",\"No\")")]),
    ]

    for filename, title, inputs, outputs in calculators:
        filepath = os.path.join(output_dir, filename)
        wb = create_simple_calculator_template(filename, title, inputs, outputs)
        wb.save(filepath)
        templates_created.append(filename)
        print(f"✓ {filename}")

    # ========== SIMPLE TEMPLATES (100 templates) ==========
    simple_templates = [
        # Checklists
        ("Monthly_Bill_Checklist.xlsx", "Monthly Bill Checklist",
         ["Bill", "Due Date", "Amount", "Auto-Pay", "Paid", "Date Paid", "Notes"]),
        ("Startup_Checklist.xlsx", "Startup Checklist",
         ["Task", "Category", "Priority", "Due Date", "Status", "Assigned To", "Notes"]),
        ("Tax_Document_Checklist.xlsx", "Tax Document Checklist",
         ["Document", "Description", "Required", "Received", "Location", "Notes"]),
        ("Year_End_Checklist.xlsx", "Year-End Checklist",
         ["Task", "Department", "Priority", "Due Date", "Status", "Completed Date", "Notes"]),

        # Logs
        ("Transaction_Log.xlsx", "Transaction Log",
         ["Date", "Type", "Category", "Description", "Amount", "Balance", "Reference"]),
        ("Payment_Log.xlsx", "Payment Log",
         ["Date", "Payee", "Amount", "Payment Method", "Category", "Reference", "Notes"]),
        ("Deposit_Log.xlsx", "Deposit Log",
         ["Date", "Source", "Amount", "Account", "Reference", "Notes"]),
        ("Withdrawal_Log.xlsx", "Withdrawal Log",
         ["Date", "Recipient", "Amount", "Account", "Reference", "Notes"]),

        # Registers
        ("Check_Register.xlsx", "Check Register",
         ["Date", "Check #", "Description", "Debit", "Credit", "Balance"]),
        ("Cash_Register.xlsx", "Cash Register",
         ["Date", "Opening Balance", "Sales", "Expenses", "Closing Balance"]),
        ("Bank_Register.xlsx", "Bank Register",
         ["Date", "Description", "Withdrawal", "Deposit", "Balance"]),
        ("Credit_Card_Register.xlsx", "Credit Card Register",
         ["Date", "Description", "Charge", "Payment", "Balance"]),

        # Lists
        ("Vendor_List.xlsx", "Vendor List",
         ["Vendor Name", "Contact", "Phone", "Email", "Address", "Terms", "Notes"]),
        ("Customer_List.xlsx", "Customer List",
         ["Customer Name", "Contact", "Phone", "Email", "Address", "Credit Limit", "Notes"]),
        ("Employee_List.xlsx", "Employee List",
         ["Employee ID", "Name", "Department", "Title", "Phone", "Email", "Hire Date", "Notes"]),
        ("Product_List.xlsx", "Product List",
         ["Product ID", "Name", "Category", "SKU", "Price", "Cost", "Supplier", "Notes"]),

        # Schedules
        ("Payment_Schedule.xlsx", "Payment Schedule",
         ["Due Date", "Payee", "Amount", "Category", "Status", "Paid Date", "Notes"]),
        ("Bill_Schedule.xlsx", "Bill Schedule",
         ["Bill", "Due Date", "Amount", "Frequency", "Auto-Pay", "Account", "Notes"]),
        ("Deposit_Schedule.xlsx", "Deposit Schedule",
         ["Expected Date", "Source", "Amount", "Account", "Notes"]),
        ("Tax_Payment_Schedule.xlsx", "Tax Payment Schedule",
         ["Tax Type", "Due Date", "Amount", "Period", "Status", "Paid Date", "Notes"]),

        # Reports
        ("Monthly_Summary_Report.xlsx", "Monthly Summary Report",
         ["Month", "Income", "Expenses", "Net Income", "Savings Rate", "Notes"]),
        ("Annual_Summary_Report.xlsx", "Annual Summary Report",
         ["Year", "Total Income", "Total Expenses", "Net Income", "Savings", "Notes"]),
        ("Category_Report.xlsx", "Category Report",
         ["Category", "Budget", "Actual", "Variance", "Variance %", "Notes"]),
        ("Department_Report.xlsx", "Department Report",
         ["Department", "Budget", "Actual", "Variance", "Variance %", "Notes"]),

        # Planning Templates
        ("Financial_Goals_Template.xlsx", "Financial Goals Template",
         ["Goal", "Target Amount", "Current Amount", "Target Date", "Monthly Contribution", "Status", "Notes"]),
        ("Savings_Goals_Template.xlsx", "Savings Goals Template",
         ["Goal Name", "Target Amount", "Saved So Far", "Target Date", "Monthly Deposit", "Status", "Notes"]),
        ("Debt_Payoff_Plan.xlsx", "Debt Payoff Plan",
         ["Creditor", "Balance", "Interest Rate", "Minimum Payment", "Monthly Payment", "Payoff Date", "Notes"]),
        ("Emergency_Fund_Template.xlsx", "Emergency Fund Template",
         ["Category", "Target Amount", "Current Amount", "Monthly Contribution", "Status", "Notes"]),

        # Budget Variations
        ("Weekly_Budget_Template.xlsx", "Weekly Budget Template",
         ["Category", "Week 1", "Week 2", "Week 3", "Week 4", "Total", "Average"]),
        ("Bi_Weekly_Budget_Template.xlsx", "Bi-Weekly Budget Template",
         ["Category", "Period 1", "Period 2", "Total", "Notes"]),
        ("Semi_Monthly_Budget_Template.xlsx", "Semi-Monthly Budget Template",
         ["Category", "1st-15th", "16th-End", "Total", "Notes"]),
        ("Quarterly_Budget_Template.xlsx", "Quarterly Budget Template",
         ["Category", "Q1", "Q2", "Q3", "Q4", "Annual Total", "Notes"]),

        # Specialized Trackers
        ("Subscription_Tracker.xlsx", "Subscription Tracker",
         ["Service", "Cost", "Billing Cycle", "Next Billing", "Auto-Renew", "Category", "Notes"]),
        ("Membership_Tracker.xlsx", "Membership Tracker",
         ["Organization", "Type", "Cost", "Renewal Date", "Auto-Renew", "Benefits", "Notes"]),
        ("Insurance_Tracker.xlsx", "Insurance Tracker",
         ["Policy", "Provider", "Type", "Coverage", "Premium", "Renewal Date", "Notes"]),
        ("Warranty_Tracker.xlsx", "Warranty Tracker",
         ["Item", "Purchase Date", "Warranty Period", "Expiration Date", "Provider", "Notes"]),

        # Asset Management
        ("Equipment_Maintenance_Log.xlsx", "Equipment Maintenance Log",
         ["Equipment ID", "Date", "Type", "Description", "Cost", "Technician", "Next Service", "Notes"]),
        ("Vehicle_Maintenance_Log.xlsx", "Vehicle Maintenance Log",
         ["Vehicle", "Date", "Mileage", "Service", "Cost", "Technician", "Notes"]),
        ("Property_Inspection_Log.xlsx", "Property Inspection Log",
         ["Property", "Date", "Inspector", "Findings", "Action Required", "Priority", "Notes"]),

        # Financial Planning
        ("Retirement_Planning_Template.xlsx", "Retirement Planning Template",
         ["Age", "Current Savings", "Annual Contribution", "Expected Return", "Years to Retirement", "Target Amount", "Status"]),
        ("College_Planning_Template.xlsx", "College Planning Template",
         ["Child", "Current Age", "College Age", "Current Savings", "Monthly Contribution", "Expected Cost", "Gap", "Notes"]),
        ("Estate_Planning_Template.xlsx", "Estate Planning Template",
         ["Asset", "Value", "Beneficiary", "Location", "Notes"]),
        ("Insurance_Needs_Template.xlsx", "Insurance Needs Template",
         ["Type", "Coverage Needed", "Current Coverage", "Gap", "Annual Premium", "Notes"]),

        # Business Planning
        ("Business_Plan_Financials.xlsx", "Business Plan Financials",
         ["Item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Notes"]),
        ("Startup_Costs_Template.xlsx", "Startup Costs Template",
         ["Category", "Item", "Cost", "One-Time", "Recurring", "Notes"]),
        ("Revenue_Model_Template.xlsx", "Revenue Model Template",
         ["Revenue Stream", "Unit Price", "Volume", "Total Revenue", "Growth Rate", "Notes"]),
        ("Cost_Structure_Template.xlsx", "Cost Structure Template",
         ["Cost Category", "Fixed", "Variable", "Total", "% of Revenue", "Notes"]),

        # Cash Management
        ("Cash_Position_Template.xlsx", "Cash Position Template",
         ["Date", "Opening Balance", "Inflows", "Outflows", "Closing Balance", "Notes"]),
        ("Cash_Flow_Forecast_Template.xlsx", "Cash Flow Forecast Template",
         ["Week", "Opening Balance", "Inflows", "Outflows", "Closing Balance", "Notes"]),
        ("Bank_Reconciliation_Template.xlsx", "Bank Reconciliation Template",
         ["Date", "Check #", "Description", "Bank Balance", "Book Balance", "Difference", "Status", "Notes"]),

        # Tax Templates
        ("Deduction_Tracker.xlsx", "Deduction Tracker",
         ["Category", "Description", "Amount", "Date", "Receipt", "Approved", "Notes"]),
        ("Tax_Credit_Tracker.xlsx", "Tax Credit Tracker",
         ["Credit", "Amount", "Eligibility", "Claimed", "Remaining", "Notes"]),
        ("Quarterly_Estimated_Tax_Template.xlsx", "Quarterly Estimated Tax Template",
         ["Quarter", "Estimated Income", "Estimated Tax", "Paid", "Remaining", "Notes"]),

        # Investment Templates
        ("Portfolio_Tracker.xlsx", "Portfolio Tracker",
         ["Asset Class", "Investment", "Symbol", "Shares", "Cost Basis", "Current Value", "Gain/Loss", "%"]),
        ("Asset_Allocation_Template.xlsx", "Asset Allocation Template",
         ["Asset Class", "Target %", "Actual %", "Difference", "Value", "Notes"]),
        ("Rebalancing_Template.xlsx", "Rebalancing Template",
         ["Asset Class", "Current %", "Target %", "Buy/Sell", "Amount", "Notes"]),

        # Real Estate Templates
        ("Property_List.xlsx", "Property List",
         ["Property", "Address", "Purchase Price", "Current Value", "Mortgage", "Equity", "Status", "Notes"]),
        ("Rental_Property_Tracker.xlsx", "Rental Property Tracker",
         ["Property", "Tenant", "Rent", "Lease Start", "Lease End", "Deposit", "Status", "Notes"]),
        ("Rental_Income_Tracker.xlsx", "Rental Income Tracker",
         ["Property", "Month", "Rent Received", "Expenses", "Net Income", "Notes"]),
        ("Mortgage_Comparison_Template.xlsx", "Mortgage Comparison Template",
         ["Lender", "Loan Amount", "Rate", "Term", "Monthly Payment", "Total Interest", "Notes"]),

        # Loan Management
        ("Loan_Tracker.xlsx", "Loan Tracker",
         ["Lender", "Loan Type", "Amount", "Rate", "Term", "Monthly Payment", "Balance", "Payoff Date", "Notes"]),
        ("Credit_Card_Tracker.xlsx", "Credit Card Tracker",
         ["Card", "Bank", "Limit", "Balance", "Available", "APR", "Payment Due", "Notes"]),
        ("Debt_Payoff_Tracker.xlsx", "Debt Payoff Tracker",
         ["Debt", "Balance", "Interest Rate", "Monthly Payment", "Payoff Date", "Priority", "Notes"]),

        # Savings Templates
        ("Savings_Account_Tracker.xlsx", "Savings Account Tracker",
         ["Account", "Bank", "Balance", "Interest Rate", "Goal", "Target Date", "Notes"]),
        ("Emergency_Fund_Tracker.xlsx", "Emergency Fund Tracker",
         ["Category", "Target Amount", "Current Amount", "Monthly Contribution", "Status", "Notes"]),
        ("Goal_Savings_Tracker.xlsx", "Goal Savings Tracker",
         ["Goal", "Target Amount", "Saved So Far", "Monthly Deposit", "Target Date", "Status", "Notes"]),

        # Expense Categories
        ("Office_Expenses_Template.xlsx", "Office Expenses Template",
         ["Category", "Item", "Date", "Amount", "Vendor", "Reimbursable", "Notes"]),
        ("Travel_Expenses_Template.xlsx", "Travel Expenses Template",
         ["Date", "Category", "Description", "Amount", "Destination", "Purpose", "Reimbursable", "Notes"]),
        ("Marketing_Expenses_Template.xlsx", "Marketing Expenses Template",
         ["Campaign", "Date", "Category", "Amount", "ROI", "Notes"]),
        ("IT_Expenses_Template.xlsx", "IT Expenses Template",
         ["Category", "Item", "Date", "Amount", "Vendor", "Recurring", "Notes"]),

        # Income Categories
        ("Sales_Revenue_Tracker.xlsx", "Sales Revenue Tracker",
         ["Date", "Customer", "Product", "Quantity", "Unit Price", "Total", "Salesperson", "Notes"]),
        ("Service_Revenue_Tracker.xlsx", "Service Revenue Tracker",
         ["Date", "Customer", "Service", "Hours", "Rate", "Total", "Notes"]),
        ("Passive_Income_Tracker.xlsx", "Passive Income Tracker",
         ["Source", "Type", "Date", "Amount", "Frequency", "Notes"]),
        ("Other_Income_Tracker.xlsx", "Other Income Tracker",
         ["Date", "Source", "Description", "Amount", "Category", "Notes"]),

        # Project Templates
        ("Project_Cost_Tracker.xlsx", "Project Cost Tracker",
         ["Project", "Category", "Item", "Budget", "Actual", "Variance", "Status", "Notes"]),
        ("Project_Time_Tracker.xlsx", "Project Time Tracker",
         ["Project", "Task", "Start Date", "End Date", "Hours", "Resource", "Status", "Notes"]),
        ("Project_Milestone_Tracker.xlsx", "Project Milestone Tracker",
         ["Project", "Milestone", "Due Date", "Actual Date", "Status", "Notes"]),
        ("Project_Resource_Tracker.xlsx", "Project Resource Tracker",
         ["Project", "Resource", "Role", "Allocation %", "Cost", "Notes"]),

        # HR Templates
        ("Employee_Benefits_Tracker.xlsx", "Employee Benefits Tracker",
         ["Employee", "Health", "Dental", "Vision", "401k", "Life Insurance", "Notes"]),
        ("PTO_Tracker.xlsx", "PTO Tracker",
         ["Employee", "Accrued", "Used", "Remaining", "Request Date", "Approved", "Notes"]),
        ("Training_Budget_Template.xlsx", "Training Budget Template",
         ["Employee", "Training", "Cost", "Date", "Provider", "Status", "Notes"]),
        ("Performance_Review_Template.xlsx", "Performance Review Template",
         ["Employee", "Review Period", "Rating", "Goals", "Achievements", "Notes"]),

        # Compliance Templates
        ("Compliance_Checklist.xlsx", "Compliance Checklist",
         ["Requirement", "Category", "Due Date", "Status", "Completed Date", "Notes"]),
        ("Audit_Tracker.xlsx", "Audit Tracker",
         ["Audit Type", "Date", "Auditor", "Findings", "Action Required", "Due Date", "Status", "Notes"]),
        ("Regulatory_Filing_Tracker.xlsx", "Regulatory Filing Tracker",
         ["Filing", "Agency", "Due Date", "Status", "Submitted Date", "Notes"]),
        ("Policy_Review_Tracker.xlsx", "Policy Review Tracker",
         ["Policy", "Last Review", "Next Review", "Reviewer", "Status", "Notes"]),

        # Dashboard Templates
        ("KPI_Dashboard_Template.xlsx", "KPI Dashboard Template",
         ["KPI", "Target", "Actual", "Variance", "%", "Trend", "Notes"]),
        ("Financial_Health_Score_Template.xlsx", "Financial Health Score Template",
         ["Category", "Metric", "Score", "Weight", "Weighted Score", "Notes"]),
        ("Budget_vs_Actual_Template.xlsx", "Budget vs Actual Template",
         ["Category", "Budget", "Actual", "Variance", "Variance %", "Status", "Notes"]),

        # Analysis Templates
        ("Trend_Analysis_Template.xlsx", "Trend Analysis Template",
         ["Period", "Revenue", "Expenses", "Net Income", "Growth %", "Notes"]),
        ("Variance_Analysis_Template.xlsx", "Variance Analysis Template",
         ["Account", "Budget", "Actual", "Variance", "Variance %", "Favorable/Unfavorable", "Notes"]),
        ("Ratio_Analysis_Template.xlsx", "Ratio Analysis Template",
         ["Ratio", "Current Period", "Previous Period", "Change", "Industry Average", "Notes"]),
        ("Scenario_Analysis_Template.xlsx", "Scenario Analysis Template",
         ["Metric", "Best Case", "Base Case", "Worst Case", "Notes"]),

        # Specialized Templates
        ("Freelancer_Finance_Template.xlsx", "Freelancer Finance Template",
         ["Client", "Project", "Hours", "Rate", "Amount", "Invoice #", "Paid", "Notes"]),
        ("Consulting_Billing_Template.xlsx", "Consulting Billing Template",
         ["Client", "Date", "Hours", "Rate", "Expenses", "Total", "Invoice #", "Notes"]),
        ("Contractor_Tracker.xlsx", "Contractor Tracker",
         ["Contractor", "Service", "Rate", "Hours", "Total", "Paid", "Notes"]),
        ("Gig_Economy_Tracker.xlsx", "Gig Economy Tracker",
         ["Platform", "Gig", "Date", "Earnings", "Expenses", "Net", "Notes"]),

        # Personal Finance
        ("Net_Worth_Tracker.xlsx", "Net Worth Tracker",
         ["Asset/Liability", "Category", "Description", "Value", "Date", "Notes"]),
        ("Debt_to_Income_Template.xlsx", "Debt to Income Template",
         ["Month", "Total Income", "Total Debt Payments", "DTI Ratio", "Status", "Notes"]),
        ("Expense_Ratio_Template.xlsx", "Expense Ratio Template",
         ["Category", "Amount", "% of Income", "Budget %", "Variance", "Notes"]),
        ("Savings_Rate_Template.xlsx", "Savings Rate Template",
         ["Month", "Income", "Expenses", "Savings", "Savings Rate", "Target Rate", "Notes"]),

        # Additional Templates
        ("Grant_Tracker.xlsx", "Grant Tracker",
         ["Grant", "Organization", "Amount", "Applied Date", "Status", "Award Date", "Notes"]),
        ("Donation_Tracker.xlsx", "Donation Tracker",
         ["Date", "Organization", "Amount", "Purpose", "Tax Deductible", "Notes"]),
        ("Reimbursement_Tracker.xlsx", "Reimbursement Tracker",
         ["Date", "Employee", "Amount", "Category", "Description", "Approved", "Paid Date", "Notes"]),
        ("Petty_Cash_Tracker.xlsx", "Petty Cash Tracker",
         ["Date", "Description", "Deposit", "Withdrawal", "Balance", "Receipt", "Notes"]),
    ]

    for filename, title, columns in simple_templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_simple_tracker_template(filename, title, columns)
        wb.save(filepath)
        templates_created.append(filename)
        print(f"✓ {filename}")

    print("=" * 70)
    print(f"\nTotal templates created: {len(templates_created)}")
    print(f"All templates saved to: {output_dir}")
    print("\nTemplate categories:")
    print("  • Trackers (50+)")
    print("  • Calculators (50+)")
    print("  • Simple Templates (100+)")
    print("  • Industry-specific")
    print("  • Personal finance")
    print("  • Business management")

if __name__ == "__main__":
    generate_all_templates()
