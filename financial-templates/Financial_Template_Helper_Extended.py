#!/usr/bin/env python3
"""
Extended Financial Excel Template Generator
Creates comprehensive English financial templates for accounting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_cash_flow_template():
    """Create Cash Flow Statement template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "CASH FLOW STATEMENT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "For the Period Ended:"
    ws['B2'] = "2026-01-31"

    # Operating Activities
    row = 4
    ws[f'A{row}'] = "OPERATING ACTIVITIES"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Net Income"
    ws[f'C{row}'] = 45000
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Adjustments to reconcile net income to net cash:"
    ws[f'A{row}'].font = Font(italic=True)
    row += 1

    adjustments = [
        ("Depreciation and Amortization", 5000),
        ("Accounts Receivable decrease", 3000),
        ("Inventory increase", -2000),
        ("Accounts Payable increase", 2500),
        ("Accrued Expenses decrease", -1000),
    ]

    for item, value in adjustments:
        ws[f'A{row}'] = "  " + item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Net Cash from Operating Activities"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C5:C11)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Investing Activities
    ws[f'A{row}'] = "INVESTING ACTIVITIES"
    ws[f'A{row}'].font = section_font
    row += 1

    investing = [
        ("Purchase of Property, Plant & Equipment", -15000),
        ("Sale of Equipment", 3000),
        ("Purchase of Investments", -10000),
        ("Sale of Investments", 5000),
    ]

    for item, value in investing:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Net Cash from Investing Activities"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C14:C17)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Financing Activities
    ws[f'A{row}'] = "FINANCING ACTIVITIES"
    ws[f'A{row}'].font = section_font
    row += 1

    financing = [
        ("Proceeds from Long-term Debt", 20000),
        ("Repayment of Long-term Debt", -5000),
        ("Issuance of Common Stock", 10000),
        ("Dividends Paid", -3000),
    ]

    for item, value in financing:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        row += 1

    ws[f'B{row}'] = "Net Cash from Financing Activities"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(C20:C23)"
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Net Change and Cash Balance
    ws[f'B{row}'] = "Net Change in Cash"
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=C12+C18+C24"
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Cash at Beginning of Period"
    ws[f'C{row}'] = 25000
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Cash at End of Period"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = "=C26+C25"
    ws[f'C{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'].number_format = currency_format

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20

    return wb

def create_accounts_payable_template():
    """Create Accounts Payable template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts Payable"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Invoice Number", "Vendor Name", "Invoice Date", "Due Date",
        "Amount", "Status", "Days Overdue", "Payment Date", "Payment Amount"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["INV-001", "ABC Supplies Inc.", "2026-01-05", "2026-02-05", 2500.00, "Unpaid", 0, "", 0],
        ["INV-002", "XYZ Services LLC", "2026-01-10", "2026-02-10", 1800.00, "Unpaid", 0, "", 0],
        ["INV-003", "Global Tech Corp", "2026-01-15", "2026-02-15", 3200.00, "Paid", 0, "2026-02-01", 3200.00],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Days overdue formula
        ws.cell(row=row, column=7).value = f'=IF(F{row}="Unpaid",MAX(0,TODAY()-D{row}),0)'
        # Payment remaining
        ws.cell(row=row, column=9).value = f'=IF(F{row}="Paid",E{row},0)'

    # Summary section
    ws['G102'] = "SUMMARY"
    ws['G102'].font = Font(bold=True)
    ws['A103'] = "Total Outstanding:"
    ws['C103'] = "=SUMIF(F2:F100,\"Unpaid\",E2:E100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Overdue:"
    ws['C104'] = "=SUMIF(G2:G100,\">0\",E2:E100)"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Total Paid:"
    ws['C105'] = "=SUM(I2:I100)"
    ws['C105'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    # Format currency
    for row in range(2, 100):
        for col in [5, 9]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_accounts_receivable_template():
    """Create Accounts Receivable template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accounts Receivable"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Invoice Number", "Customer Name", "Invoice Date", "Due Date",
        "Amount", "Amount Paid", "Balance Due", "Days Outstanding", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["INV-1001", "Acme Corp", "2026-01-05", "2026-02-05", 5000.00, 2000.00, 3000.00, 0, "Partial"],
        ["INV-1002", "Beta Industries", "2026-01-10", "2026-02-10", 3500.00, 0.00, 3500.00, 0, "Unpaid"],
        ["INV-1003", "Gamma Solutions", "2026-01-15", "2026-02-15", 4200.00, 4200.00, 0.00, 0, "Paid"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Balance due
        ws.cell(row=row, column=7).value = f'=E{row}-F{row}'
        # Days outstanding
        ws.cell(row=row, column=8).value = f'=IF(G{row}>0,TODAY()-D{row},0)'
        # Status based on days
        ws.cell(row=row, column=9).value = f'=IF(G{row}=0,"Paid",IF(H{row}>30,"Overdue","Current"))'

    # Summary section
    ws['G102'] = "SUMMARY"
    ws['G102'].font = Font(bold=True)
    ws['A103'] = "Total Outstanding:"
    ws['C103'] = "=SUM(G2:G100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Over 30 Days:"
    ws['C104'] = "=SUMIFS(G2:G100,G2:G100,\">0\",H2:H100,\">30\")"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Total Collected:"
    ws['C105'] = "=SUM(F2:F100)"
    ws['C105'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10

    # Format currency
    for row in range(2, 100):
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_payroll_template():
    """Create Payroll Calculator template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll Calculator"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Employee ID", "Employee Name", "Department", "Gross Pay",
        "Federal Tax", "State Tax", "Social Security", "Medicare",
        "401(k)", "Health Insurance", "Other Deductions", "Net Pay"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["EMP001", "John Smith", "Sales", 5000.00, 500.00, 250.00, 310.00, 72.50, 200.00, 150.00, 0, 3517.50],
        ["EMP002", "Jane Doe", "Marketing", 4800.00, 480.00, 240.00, 297.60, 69.60, 192.00, 150.00, 50, 3320.80],
        ["EMP003", "Bob Johnson", "IT", 5500.00, 550.00, 275.00, 341.00, 79.75, 220.00, 150.00, 0, 3884.25],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Federal Tax (10%)
        ws.cell(row=row, column=5).value = f'=D{row}*0.10'
        # State Tax (5%)
        ws.cell(row=row, column=6).value = f'=D{row}*0.05'
        # Social Security (6.2%)
        ws.cell(row=row, column=7).value = f'=D{row}*0.062'
        # Medicare (1.45%)
        ws.cell(row=row, column=8).value = f'=D{row}*0.0145'
        # 401(k) (4%)
        ws.cell(row=row, column=9).value = f'=D{row}*0.04'
        # Net Pay
        ws.cell(row=row, column=12).value = f'=D{row}-SUM(E{row}:K{row})'

    # Summary section
    ws['J102'] = "PAYROLL SUMMARY"
    ws['J102'].font = Font(bold=True, size=12)
    ws['A103'] = "Total Gross Pay:"
    ws['C103'] = "=SUM(D2:D100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Federal Tax:"
    ws['C104'] = "=SUM(E2:E100)"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Total State Tax:"
    ws['C105'] = "=SUM(F2:F100)"
    ws['C105'].number_format = '#,##0.00'
    ws['A106'] = "Total Net Pay:"
    ws['C106'] = "=SUM(L2:L100)"
    ws['C106'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    for col in range(4, 13):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Format currency
    for row in range(2, 100):
        for col in range(4, 13):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_inventory_template():
    """Create Inventory Management template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Management"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "SKU", "Product Name", "Category", "Unit Cost",
        "Selling Price", "Quantity on Hand", "Reorder Level",
        "Total Value", "Potential Revenue", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["SKU001", "Laptop Computer", "Electronics", 500.00, 799.99, 25, 10, 12500.00, 19999.75, "OK"],
        ["SKU002", "Office Chair", "Furniture", 150.00, 299.99, 5, 15, 750.00, 1499.95, "Reorder"],
        ["SKU003", "Wireless Mouse", "Electronics", 25.00, 49.99, 100, 20, 2500.00, 4999.00, "OK"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Total Value
        ws.cell(row=row, column=8).value = f'=D{row}*G{row}'
        # Potential Revenue
        ws.cell(row=row, column=9).value = f'=E{row}*G{row}'
        # Status
        ws.cell(row=row, column=10).value = f'=IF(G{row}<H{row},"Reorder","OK")'

    # Summary section
    ws['I102'] = "INVENTORY SUMMARY"
    ws['I102'].font = Font(bold=True, size=12)
    ws['A103'] = "Total Inventory Value:"
    ws['C103'] = "=SUM(H2:H100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Potential Revenue:"
    ws['C104'] = "=SUM(I2:I100)"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Items to Reorder:"
    ws['C105'] = "=COUNTIF(J2:J100,\"Reorder\")"

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 10

    # Format currency
    for row in range(2, 100):
        for col in [4, 5, 8, 9]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_expense_tracker_template():
    """Create Expense Tracker template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Tracker"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Date", "Category", "Description", "Vendor",
        "Amount", "Payment Method", "Receipt", "Notes", "Reimbursable"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["2026-01-05", "Office Supplies", "Paper and pens", "Staples", 45.99, "Credit Card", "Yes", "Monthly supplies", "No"],
        ["2026-01-10", "Travel", "Client meeting - lunch", "Restaurant", 125.50, "Cash", "Yes", "Client expense", "Yes"],
        ["2026-01-15", "Software", "Cloud storage subscription", "Tech Corp", 29.99, "Credit Card", "No", "Monthly", "No"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Summary section
    ws['H102'] = "EXPENSE SUMMARY"
    ws['H102'].font = Font(bold=True, size=12)

    # Category breakdown
    categories = ["Office Supplies", "Travel", "Software", "Marketing", "Utilities", "Rent", "Other"]
    row = 104
    ws[f'A{row}'] = "By Category:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for cat in categories:
        ws[f'A{row}'] = cat
        ws[f'C{row}'] = f'=SUMIF(B2:B100,"{cat}",E2:E100)'
        ws[f'C{row}'].number_format = '#,##0.00'
        row += 1

    # Totals
    ws[f'A{row}'] = "TOTAL EXPENSES:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = "=SUM(E2:E100)"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'C{row}'].number_format = '#,##0.00'
    row += 1

    ws[f'A{row}'] = "Reimbursable:"
    ws[f'C{row}'] = "=SUMIFS(E2:E100,I2:I100,\"Yes\")"
    ws[f'C{row}'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 12

    # Format currency
    for row in range(2, 100):
        ws.cell(row=row, column=5).number_format = '#,##0.00'

    return wb

def create_invoice_template():
    """Create Professional Invoice template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Company Information
    ws['A1'] = "YOUR COMPANY NAME"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "123 Business Street"
    ws['A3'] = "City, State 12345"
    ws['A4'] = "Phone: (555) 123-4567"
    ws['A5'] = "Email: info@yourcompany.com"

    # Invoice Details
    ws['G1'] = "INVOICE"
    ws['G1'].font = Font(bold=True, size=20)
    ws['G1'].alignment = Alignment(horizontal='right')

    ws['G3'] = "Invoice #:"
    ws['H3'] = "INV-001"
    ws['G4'] = "Date:"
    ws['H4'] = "2026-01-31"
    ws['G5'] = "Due Date:"
    ws['H5'] = "2026-02-28"

    # Bill To
    ws['A8'] = "BILL TO:"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = "Client Company Name"
    ws['A10'] = "456 Client Avenue"
    ws['A11'] = "City, State 67890"

    # Line Items
    ws['A14'] = "Description"
    ws['A14'].font = Font(bold=True)
    ws['F14'] = "Quantity"
    ws['F14'].font = Font(bold=True)
    ws['G14'] = "Unit Price"
    ws['G14'].font = Font(bold=True)
    ws['H14'] = "Amount"
    ws['H14'].font = Font(bold=True)

    # Sample line items
    line_items = [
        ["Product/Service 1", 5, 100.00],
        ["Product/Service 2", 3, 250.00],
        ["Product/Service 3", 2, 75.00],
    ]

    for row_idx, (desc, qty, price) in enumerate(line_items, start=15):
        ws[f'A{row_idx}'] = desc
        ws[f'F{row_idx}'] = qty
        ws[f'G{row_idx}'] = price
        ws[f'G{row_idx}'].number_format = '#,##0.00'
        ws[f'H{row_idx}'] = f'=F{row_idx}*G{row_idx}'
        ws[f'H{row_idx}'].number_format = '#,##0.00'

    # Totals
    ws['F19'] = "Subtotal:"
    ws['F19'].font = Font(bold=True)
    ws['H19'] = "=SUM(H15:H17)"
    ws['H19'].font = Font(bold=True)
    ws['H19'].number_format = '#,##0.00'

    ws['F20'] = "Tax (10%):"
    ws['H20'] = "=H19*0.10"
    ws['H20'].number_format = '#,##0.00'

    ws['F21'] = "Total:"
    ws['F21'].font = Font(bold=True, size=12)
    ws['H21'] = "=H19+H20"
    ws['H21'].font = Font(bold=True, size=12)
    ws['H21'].number_format = '#,##0.00'

    # Notes
    ws['A24'] = "Notes:"
    ws['A24'].font = Font(bold=True)
    ws['A25'] = "Payment is due within 30 days."
    ws['A26'] = "Thank you for your business!"

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    return wb

def create_project_budget_template():
    """Create Project Budget template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Budget"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Cost Category", "Item Description", "Estimated Cost",
        "Actual Cost", "Variance", "Variance %", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["Personnel", "Project Manager", 8000.00, 7500.00, -500.00, -0.0625, "Under Budget"],
        ["Personnel", "Developers (2)", 20000.00, 22000.00, 2000.00, 0.10, "Over Budget"],
        ["Equipment", "Computers", 5000.00, 4800.00, -200.00, -0.04, "Under Budget"],
        ["Software", "Licenses", 2000.00, 2000.00, 0.00, 0.00, "On Budget"],
        ["Marketing", "Advertising", 3000.00, 3500.00, 500.00, 0.1667, "Over Budget"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Variance
        ws.cell(row=row, column=5).value = f'=D{row}-C{row}'
        # Variance %
        ws.cell(row=row, column=6).value = f'=IF(C{row}<>0,(D{row}-C{row})/C{row},0)'
        # Status
        ws.cell(row=row, column=7).value = f'=IF(E{row}>0,"Over Budget",IF(E{row}<0,"Under Budget","On Budget"))'

    # Summary section
    ws['G102'] = "PROJECT BUDGET SUMMARY"
    ws['G102'].font = Font(bold=True, size=12)
    ws['A103'] = "Total Estimated:"
    ws['C103'] = "=SUM(C2:C100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Actual:"
    ws['C104'] = "=SUM(D2:D100)"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Total Variance:"
    ws['C105'] = "=SUM(E2:E100)"
    ws['C105'].number_format = '#,##0.00'
    ws['A106'] = "Variance %:"
    ws['C106'] = "=IF(C103<>0,E105/C103,0)"
    ws['C106'].number_format = '0.00%'

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # Format currency and percentage
    for row in range(2, 100):
        for col in [3, 4, 5]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = '0.00%'

    return wb

def create_break_even_template():
    """Create Break-Even Analysis template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Break-Even Analysis"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    number_format = '#,##0.00'

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "BREAK-EVEN ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Input Section
    row = 3
    ws[f'A{row}'] = "INPUT PARAMETERS"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Fixed Costs:"
    ws[f'C{row}'] = 10000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Variable Cost per Unit:"
    ws[f'C{row}'] = 25.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Selling Price per Unit:"
    ws[f'C{row}'] = 50.00
    ws[f'C{row}'].number_format = currency_format
    row += 2

    # Calculations
    ws[f'A{row}'] = "CALCULATIONS"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Contribution Margin per Unit:"
    ws[f'C{row}'] = "=C6-C5"
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Contribution Margin Ratio:"
    ws[f'C{row}'] = "=C8/C6"
    ws[f'C{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'] = "Break-Even Point (Units):"
    ws[f'C{row}'] = "=C4/C8"
    ws[f'C{row}'].number_format = number_format
    row += 1

    ws[f'A{row}'] = "Break-Even Point (Dollars):"
    ws[f'C{row}'] = "=C10*C6"
    ws[f'C{row}'].number_format = currency_format
    # Target Profit Analysis
    row += 2
    ws[f'A{row}'] = "TARGET PROFIT ANALYSIS"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Target Profit:"
    ws[f'C{row}'] = 5000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Units Needed for Target Profit:"
    ws[f'C{row}'] = "=(C4+C13)/C8"
    ws[f'C{row}'].number_format = number_format
    row += 1

    ws[f'A{row}'] = "Revenue Needed for Target Profit:"
    ws[f'C{row}'] = "=C14*C6"
    ws[f'C{row}'].number_format = currency_format

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20

    return wb

def create_sales_forecast_template():
    """Create Sales Forecast template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Forecast"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    months = ["Product", "Jan", "Feb", "Mar", "Q1 Total", "Apr", "May", "Jun", "Q2 Total", 
              "Jul", "Aug", "Sep", "Q3 Total", "Oct", "Nov", "Dec", "Q4 Total", "Year Total"]

    for col, month in enumerate(months, start=1):
        cell = ws.cell(row=1, column=col, value=month)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    products = [
        ["Product A", 10000, 11000, 12000, 0, 13000, 14000, 15000, 0, 16000, 17000, 18000, 0, 19000, 20000, 21000, 0, 0],
        ["Product B", 8000, 8500, 9000, 0, 9500, 10000, 10500, 0, 11000, 11500, 12000, 0, 12500, 13000, 13500, 0, 0],
        ["Product C", 5000, 5500, 6000, 0, 6500, 7000, 7500, 0, 8000, 8500, 9000, 0, 9500, 10000, 10500, 0, 0],
    ]

    for row_idx, row_data in enumerate(products, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas for quarterly and yearly totals
    for row in range(2, 5):
        # Q1
        ws.cell(row=row, column=5).value = f'=SUM(B{row}:D{row})'
        # Q2
        ws.cell(row=row, column=9).value = f'=SUM(F{row}:H{row})'
        # Q3
        ws.cell(row=row, column=13).value = f'=SUM(J{row}:L{row})'
        # Q4
        ws.cell(row=row, column=17).value = f'=SUM(N{row}:P{row})'
        # Year Total
        ws.cell(row=row, column=18).value = f'=SUM(B{row}:P{row})'

    # Total row
    ws['A6'] = "TOTAL"
    ws['A6'].font = Font(bold=True)
    for col in range(2, 19):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}6'].value = f'=SUM({col_letter}2:{col_letter}5)'
        ws[f'{col_letter}6'].font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, 19):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Format as currency
    for row in range(2, 7):
        for col in range(2, 19):
            ws.cell(row=row, column=col).number_format = '#,##0'

    return wb

def create_variance_analysis_template():
    """Create Variance Analysis template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variance Analysis"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Account", "Budget", "Actual", "Variance", "Variance %", 
        "Favorable/Unfavorable", "Notes"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["Revenue", 100000, 105000, 5000, 0.05, "Favorable", ""],
        ["Cost of Goods Sold", 40000, 38000, -2000, -0.05, "Favorable", ""],
        ["Salaries", 25000, 26000, 1000, 0.04, "Unfavorable", ""],
        ["Rent", 5000, 5000, 0, 0.00, "Neutral", ""],
        ["Utilities", 2000, 2500, 500, 0.25, "Unfavorable", ""],
        ["Marketing", 3000, 3500, 500, 0.1667, "Unfavorable", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 50):
        # Variance
        ws.cell(row=row, column=4).value = f'=C{row}-B{row}'
        # Variance %
        ws.cell(row=row, column=5).value = f'=IF(B{row}<>0,(C{row}-B{row})/B{row},0)'
        # Favorable/Unfavorable
        ws.cell(row=row, column=6).value = f'=IF(A{row}="Revenue",IF(D{row}>0,"Favorable","Unfavorable"),IF(D{row}<0,"Favorable","Unfavorable"))'

    # Summary section
    ws['F52'] = "VARIANCE SUMMARY"
    ws['F52'].font = Font(bold=True, size=12)
    ws['A53'] = "Total Budget:"
    ws['C53'] = "=SUM(B2:B50)"
    ws['C53'].number_format = '#,##0.00'
    ws['A54'] = "Total Actual:"
    ws['C54'] = "=SUM(C2:C50)"
    ws['C54'].number_format = '#,##0.00'
    ws['A55'] = "Total Variance:"
    ws['C55'] = "=SUM(D2:D50)"
    ws['C55'].number_format = '#,##0.00'
    ws['A56'] = "Net Income Variance:"
    ws['C56'] = "=D2-SUM(D3:D50)"
    ws['C56'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30

    # Format currency and percentage
    for row in range(2, 50):
        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '0.00%'

    return wb

def main():
    """Generate all extended financial templates"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"

    print("Generating Extended Financial Excel Templates...")
    print("=" * 60)

    templates = [
        ("Cash_Flow_Statement.xlsx", create_cash_flow_template),
        ("Accounts_Payable.xlsx", create_accounts_payable_template),
        ("Accounts_Receivable.xlsx", create_accounts_receivable_template),
        ("Payroll_Calculator.xlsx", create_payroll_template),
        ("Inventory_Management.xlsx", create_inventory_template),
        ("Expense_Tracker.xlsx", create_expense_tracker_template),
        ("Professional_Invoice.xlsx", create_invoice_template),
        ("Project_Budget.xlsx", create_project_budget_template),
        ("Break_Even_Analysis.xlsx", create_break_even_template),
        ("Sales_Forecast.xlsx", create_sales_forecast_template),
        ("Variance_Analysis.xlsx", create_variance_analysis_template),
    ]

    for filename, create_func in templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_func()
        wb.save(filepath)
        print(f"✓ Created: {filename}")

    print("=" * 60)
    print(f"\nAll templates saved to: {output_dir}")
    print("\nNew templates include:")
    print("  • Cash Flow Statement")
    print("  • Accounts Payable/Receivable")
    print("  • Payroll Calculator")
    print("  • Inventory Management")
    print("  • Expense Tracker")
    print("  • Professional Invoice")
    print("  • Project Budget")
    print("  • Break-Even Analysis")
    print("  • Sales Forecast")
    print("  • Variance Analysis")

if __name__ == "__main__":
    main()
