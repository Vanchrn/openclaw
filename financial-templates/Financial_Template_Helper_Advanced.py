#!/usr/bin/env python3
"""
Advanced Financial Excel Template Generator
Creates specialized English financial templates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_depreciation_schedule_template():
    """Create Depreciation Schedule template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depreciation Schedule"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Asset ID", "Asset Name", "Purchase Date", "Cost", "Salvage Value",
        "Useful Life (Years)", "Depreciation Method", "Annual Depreciation",
        "Accumulated Depreciation", "Book Value", "Status"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["AST001", "Office Building", "2020-01-01", 500000.00, 50000.00, 40, "Straight-Line", 11250.00, 67500.00, 432500.00, "Active"],
        ["AST002", "Delivery Truck", "2021-06-01", 45000.00, 5000.00, 5, "Double Declining", 9000.00, 18000.00, 27000.00, "Active"],
        ["AST003", "Computer Equipment", "2022-01-01", 15000.00, 0.00, 5, "Straight-Line", 3000.00, 6000.00, 9000.00, "Active"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Add formulas
    for row in range(2, 100):
        # Annual Depreciation (Straight-Line)
        ws.cell(row=row, column=8).value = f'=IF(G{row}="Straight-Line",(D{row}-E{row})/F{row},0)'
        # Accumulated Depreciation
        ws.cell(row=row, column=9).value = f'=H{row}*(YEAR(TODAY())-YEAR(C{row}))'
        # Book Value
        ws.cell(row=row, column=10).value = f'=D{row}-I{row}'

    # Summary section
    ws['J102'] = "DEPRECIATION SUMMARY"
    ws['J102'].font = Font(bold=True, size=12)
    ws['A103'] = "Total Asset Cost:"
    ws['C103'] = "=SUM(D2:D100)"
    ws['C103'].number_format = '#,##0.00'
    ws['A104'] = "Total Accumulated Depreciation:"
    ws['C104'] = "=SUM(I2:I100)"
    ws['C104'].number_format = '#,##0.00'
    ws['A105'] = "Total Book Value:"
    ws['C105'] = "=SUM(J2:J100)"
    ws['C105'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10

    # Format currency
    for row in range(2, 100):
        for col in [4, 5, 8, 9, 10]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    return wb

def create_loan_amortization_template():
    """Create Loan Amortization Schedule template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Amortization"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Period", "Payment Date", "Beginning Balance", "Payment",
        "Principal", "Interest", "Ending Balance", "Cumulative Interest"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Loan parameters
    ws['K1'] = "Loan Parameters:"
    ws['K2'] = "Loan Amount:"
    ws['L2'] = 100000.00
    ws['L2'].number_format = '#,##0.00'
    ws['K3'] = "Annual Interest Rate:"
    ws['L3'] = 0.06
    ws['L3'].number_format = '0.00%'
    ws['K4'] = "Loan Term (Years):"
    ws['L4'] = 30
    ws['K5'] = "Monthly Payment:"
    ws['L5'] = "=PMT(L3/12, L4*12, -L2)"
    ws['L5'].number_format = '#,##0.00'

    # Generate schedule
    monthly_payment = 599.55  # Calculated from PMT(6%/12, 360, -100000)
    balance = 100000.00
    start_date = "2026-01-01"

    for period in range(1, 361):
        row = period + 1
        interest = balance * 0.005  # 6%/12
        principal = monthly_payment - interest
        balance = balance - principal

        if balance < 0:
            principal = principal + balance
            balance = 0

        ws.cell(row=row, column=1).value = period
        ws.cell(row=row, column=2).value = start_date
        ws.cell(row=row, column=3).value = 100000.00 if period == 1 else ws.cell(row=row-1, column=7).value
        ws.cell(row=row, column=4).value = monthly_payment
        ws.cell(row=row, column=5).value = principal
        ws.cell(row=row, column=6).value = interest
        ws.cell(row=row, column=7).value = balance
        ws.cell(row=row, column=8).value = f'=H{row-1}+F{row}' if row > 2 else interest

    # Format currency
    for row in range(2, 362):
        for col in [3, 4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    return wb

def create_chart_of_accounts_template():
    """Create Chart of Accounts template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Account Number", "Account Name", "Account Type",
        "Normal Balance", "Subtype", "Description", "Active"
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample data
    sample_data = [
        ["1000", "Cash", "Asset", "Debit", "Current Asset", "Cash on hand and in bank", "Yes"],
        ["1100", "Accounts Receivable", "Asset", "Debit", "Current Asset", "Money owed by customers", "Yes"],
        ["1200", "Inventory", "Asset", "Debit", "Current Asset", "Goods for sale", "Yes"],
        ["1500", "Property, Plant & Equipment", "Asset", "Debit", "Non-Current Asset", "Long-term assets", "Yes"],
        ["2000", "Accounts Payable", "Liability", "Credit", "Current Liability", "Money owed to vendors", "Yes"],
        ["2100", "Accrued Expenses", "Liability", "Credit", "Current Liability", "Expenses incurred but not paid", "Yes"],
        ["2500", "Long-term Debt", "Liability", "Credit", "Non-Current Liability", "Long-term obligations", "Yes"],
        ["3000", "Owner's Capital", "Equity", "Credit", "Owner's Equity", "Owner's investment", "Yes"],
        ["3100", "Retained Earnings", "Equity", "Credit", "Retained Earnings", "Accumulated profits", "Yes"],
        ["4000", "Sales Revenue", "Revenue", "Credit", "Operating Revenue", "Income from sales", "Yes"],
        ["4100", "Service Revenue", "Revenue", "Credit", "Operating Revenue", "Income from services", "Yes"],
        ["5000", "Cost of Goods Sold", "Expense", "Debit", "Cost of Sales", "Direct costs of goods", "Yes"],
        ["6000", "Salaries Expense", "Expense", "Debit", "Operating Expense", "Employee compensation", "Yes"],
        ["6100", "Rent Expense", "Expense", "Debit", "Operating Expense", "Facility costs", "Yes"],
        ["6200", "Utilities Expense", "Expense", "Debit", "Operating Expense", "Utility costs", "Yes"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    # Summary section
    ws['F102'] = "CHART OF ACCOUNTS SUMMARY"
    ws['F102'].font = Font(bold=True, size=12)
    ws['A103'] = "Total Active Accounts:"
    ws['C103'] = "=COUNTIF(G2:G100,\"Yes\")"
    ws['A104'] = "Asset Accounts:"
    ws['C104'] = "=COUNTIFS(C2:C100,\"Asset\",G2:G100,\"Yes\")"
    ws['A105'] = "Liability Accounts:"
    ws['C105'] = "=COUNTIFS(C2:C100,\"Liability\",G2:G100,\"Yes\")"
    ws['A106'] = "Equity Accounts:"
    ws['C106'] = "=COUNTIFS(C2:C100,\"Equity\",G2:G100,\"Yes\")"
    ws['A107'] = "Revenue Accounts:"
    ws['C107'] = "=COUNTIFS(C2:C100,\"Revenue\",G2:G100,\"Yes\")"
    ws['A108'] = "Expense Accounts:"
    ws['C108'] = "=COUNTIFS(C2:C100,\"Expense\",G2:G100,\"Yes\")"

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 10

    return wb

def create_monthly_financial_report_template():
    """Create Monthly Financial Report template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Financial Report"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "MONTHLY FINANCIAL REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "Company:"
    ws['B2'] = "Your Company Name"
    ws['A3'] = "Report Period:"
    ws['B3'] = "January 2026"

    # Key Metrics
    row = 5
    ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
    ws[f'A{row}'].font = section_font
    row += 1

    metrics = [
        ("Total Revenue", 150000.00),
        ("Gross Profit", 90000.00),
        ("Operating Income", 35000.00),
        ("Net Income", 25000.00),
        ("Gross Margin %", 0.60),
        ("Operating Margin %", 0.2333),
        ("Net Profit Margin %", 0.1667),
        ("Current Ratio", 2.71),
        ("Quick Ratio", 1.88),
        ("Debt-to-Equity", 0.84),
    ]

    for metric, value in metrics:
        ws[f'A{row}'] = metric
        ws[f'C{row}'] = value
        if "%" in metric or "Ratio" in metric:
            ws[f'C{row}'].number_format = percent_format
        else:
            ws[f'C{row}'].number_format = currency_format
        row += 1

    # Comparison with Previous Month
    row += 1
    ws[f'A{row}'] = "MONTH-OVER-MONTH COMPARISON"
    ws[f'A{row}'].font = section_font
    row += 1

    ws['A' + str(row)] = "Metric"
    ws['B' + str(row)] = "Current Month"
    ws['C' + str(row)] = "Previous Month"
    ws['D' + str(row)] = "Change"
    ws['E' + str(row)] = "Change %"
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(bold=True)
    row += 1

    comparisons = [
        ("Revenue", 150000.00, 140000.00),
        ("Expenses", 125000.00, 130000.00),
        ("Net Income", 25000.00, 10000.00),
    ]

    for metric, current, previous in comparisons:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = current
        ws[f'C{row}'] = previous
        ws[f'D{row}'] = current - previous
        ws[f'E{row}'] = (current - previous) / previous if previous != 0 else 0
        ws[f'B{row}'].number_format = currency_format
        ws[f'C{row}'].number_format = currency_format
        ws[f'D{row}'].number_format = currency_format
        ws[f'E{row}'].number_format = percent_format
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12

    return wb

def create_tax_worksheet_template():
    """Create Tax Worksheet template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Worksheet"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "INCOME TAX WORKSHEET"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = "Tax Year:"
    ws['B2'] = "2026"

    # Income Section
    row = 4
    ws[f'A{row}'] = "INCOME"
    ws[f'A{row}'].font = section_font
    row += 1

    income_items = [
        ("Gross Receipts or Sales", 500000.00),
        ("Returns and Allowances", -10000.00),
        ("Cost of Goods Sold", -200000.00),
        ("Gross Profit", 290000.00),
        ("Other Income", 15000.00),
        ("TOTAL INCOME", 305000.00),
    ]

    for item, value in income_items:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Deductions Section
    row += 1
    ws[f'A{row}'] = "DEDUCTIONS"
    ws[f'A{row}'].font = section_font
    row += 1

    deductions = [
        ("Salaries and Wages", 80000.00),
        ("Rent", 24000.00),
        ("Utilities", 6000.00),
        ("Depreciation", 10000.00),
        ("Other Business Expenses", 35000.00),
        ("TOTAL DEDUCTIONS", 155000.00),
    ]

    for item, value in deductions:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "TOTAL" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Tax Calculation
    row += 1
    ws[f'A{row}'] = "TAX CALCULATION"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Taxable Income:"
    ws[f'C{row}'] = "=C12-C20"
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Tax Rate:"
    ws[f'C{row}'] = 0.21
    ws[f'C{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'] = "Income Tax Liability:"
    ws[f'C{row}'] = "=C22*C23"
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Estimated Tax Payments:"
    ws[f'C{row}'] = 50000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Tax Due (Refund):"
    ws[f'C{row}'] = "=C24-C25"
    ws[f'C{row}'].number_format = currency_format
    ws[f'C{row}'].font = Font(bold=True, size=12)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20

    return wb

def create_capital_budgeting_template():
    """Create Capital Budgeting Analysis template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capital Budgeting"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    percent_format = '0.00%'

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "CAPITAL BUDGETING ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Project Information
    row = 3
    ws[f'A{row}'] = "PROJECT INFORMATION"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Initial Investment:"
    ws[f'C{row}'] = -100000.00
    ws[f'C{row}'].number_format = currency_format
    row += 1

    ws[f'A{row}'] = "Discount Rate (WACC):"
    ws[f'C{row}'] = 0.10
    ws[f'C{row}'].number_format = percent_format
    row += 1

    ws[f'A{row}'] = "Project Life (Years):"
    ws[f'C{row}'] = 5
    row += 2

    # Cash Flow Projections
    ws[f'A{row}'] = "CASH FLOW PROJECTIONS"
    ws[f'A{row}'].font = section_font
    row += 1

    ws['A' + str(row)] = "Year"
    ws['B' + str(row)] = "Cash Flow"
    ws['C' + str(row)] = "PV Factor"
    ws['D' + str(row)] = "Present Value"
    ws['E' + str(row)] = "Cumulative PV"
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = Font(bold=True)
    row += 1

    cash_flows = [
        (0, -100000.00),
        (1, 30000.00),
        (2, 35000.00),
        (3, 40000.00),
        (4, 35000.00),
        (5, 30000.00),
    ]

    for year, cf in cash_flows:
        ws[f'A{row}'] = year
        ws[f'B{row}'] = cf
        ws[f'B{row}'].number_format = currency_format
        if year == 0:
            ws[f'C{row}'] = 1.00
        else:
            ws[f'C{row}'] = f'=1/POWER(1+$C$5,{year})'
        ws[f'C{row}'].number_format = '0.0000'
        ws[f'D{row}'] = f'=B{row}*C{row}'
        ws[f'D{row}'].number_format = currency_format
        if year == 0:
            ws[f'E{row}'] = f'=D{row}'
        else:
            ws[f'E{row}'] = f'=E{row-1}+D{row}'
        ws[f'E{row}'].number_format = currency_format
        row += 1

    # Investment Metrics
    row += 1
    ws[f'A{row}'] = "INVESTMENT METRICS"
    ws[f'A{row}'].font = section_font
    row += 1

    metrics = [
        ("Net Present Value (NPV)", "=SUM(D8:D13)"),
        ("Internal Rate of Return (IRR)", "=IRR(B8:B13,0.1)"),
        ("Payback Period (Years)", "=MATCH(0,E8:E13,-1)"),
        ("Profitability Index", "=SUM(D9:D13)/ABS(D8)"),
    ]

    for metric, formula in metrics:
        ws[f'A{row}'] = metric
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = currency_format if "NPV" in metric else '0.00%'
        row += 1

    # Decision
    row += 1
    ws[f'A{row}'] = "INVESTMENT DECISION:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'C{row}'] = '=IF(C15>0,"ACCEPT","REJECT")'
    ws[f'C{row}'].font = Font(bold=True, size=12)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    return wb

def create_working_capital_template():
    """Create Working Capital Management template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Working Capital"

    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    currency_format = '#,##0.00'
    number_format = '#,##0.00'

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "WORKING CAPITAL MANAGEMENT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Current Assets
    row = 3
    ws[f'A{row}'] = "CURRENT ASSETS"
    ws[f'A{row}'].font = section_font
    row += 1

    current_assets = [
        ("Cash and Cash Equivalents", 50000.00),
        ("Accounts Receivable", 80000.00),
        ("Inventory", 60000.00),
        ("Prepaid Expenses", 10000.00),
        ("Total Current Assets", 200000.00),
    ]

    for item, value in current_assets:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "Total" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Current Liabilities
    row += 1
    ws[f'A{row}'] = "CURRENT LIABILITIES"
    ws[f'A{row}'].font = section_font
    row += 1

    current_liab = [
        ("Accounts Payable", 45000.00),
        ("Short-term Debt", 20000.00),
        ("Accrued Expenses", 15000.00),
        ("Total Current Liabilities", 80000.00),
    ]

    for item, value in current_liab:
        ws[f'A{row}'] = item
        ws[f'C{row}'] = value
        ws[f'C{row}'].number_format = currency_format
        if "Total" in item:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'].font = Font(bold=True)
        row += 1

    # Working Capital Metrics
    row += 1
    ws[f'A{row}'] = "WORKING CAPITAL METRICS"
    ws[f'A{row}'].font = section_font
    row += 1

    metrics = [
        ("Net Working Capital", "=C9-C14", currency_format),
        ("Current Ratio", "=C9/C14", number_format),
        ("Quick Ratio", "=(C9-C6)/C14", number_format),
        ("Cash Ratio", "=C4/C14", number_format),
    ]

    for metric, formula, fmt in metrics:
        ws[f'A{row}'] = metric
        ws[f'C{row}'] = formula
        ws[f'C{row}'].number_format = fmt
        row += 1

    # Cash Conversion Cycle
    row += 1
    ws[f'A{row}'] = "CASH CONVERSION CYCLE"
    ws[f'A{row}'].font = section_font
    row += 1

    ws[f'A{row}'] = "Days Sales Outstanding (DSO):"
    ws[f'C{row}'] = 45.00
    ws[f'C{row}'].number_format = '#,##0.00'
    row += 1

    ws[f'A{row}'] = "Days Inventory Outstanding (DIO):"
    ws[f'C{row}'] = 60.00
    ws[f'C{row}'].number_format = '#,##0.00'
    row += 1

    ws[f'A{row}'] = "Days Payable Outstanding (DPO):"
    ws[f'C{row}'] = 30.00
    ws[f'C{row}'].number_format = '#,##0.00'
    row += 1

    ws[f'A{row}'] = "Cash Conversion Cycle:"
    ws[f'C{row}'] = "=C20+C21-C22"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    return wb

def main():
    """Generate all advanced financial templates"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"

    print("Generating Advanced Financial Excel Templates...")
    print("=" * 60)

    templates = [
        ("Depreciation_Schedule.xlsx", create_depreciation_schedule_template),
        ("Loan_Amortization.xlsx", create_loan_amortization_template),
        ("Chart_of_Accounts.xlsx", create_chart_of_accounts_template),
        ("Monthly_Financial_Report.xlsx", create_monthly_financial_report_template),
        ("Tax_Worksheet.xlsx", create_tax_worksheet_template),
        ("Capital_Budgeting.xlsx", create_capital_budgeting_template),
        ("Working_Capital.xlsx", create_working_capital_template),
    ]

    for filename, create_func in templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_func()
        wb.save(filepath)
        print(f"✓ Created: {filename}")

    print("=" * 60)
    print(f"\nAll templates saved to: {output_dir}")
    print("\nAdvanced templates include:")
    print("  • Depreciation Schedule")
    print("  • Loan Amortization")
    print("  • Chart of Accounts")
    print("  • Monthly Financial Report")
    print("  • Tax Worksheet")
    print("  • Capital Budgeting Analysis")
    print("  • Working Capital Management")

if __name__ == "__main__":
    main()
