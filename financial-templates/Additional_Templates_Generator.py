#!/usr/bin/env python3
"""
Additional Financial Templates Generator
Creates the remaining templates to reach 200+
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_simple_tracker_template(name, title, columns):
    """Create a simple tracker template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name.replace("_", " ").title()

    # Title
    ws.merge_cells('A1:Z1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Add some empty rows
    for row in range(4, 25):
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = border

    # Auto-fit columns
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return wb

def generate_additional_templates():
    """Generate additional templates to reach 200+"""
    output_dir = "/home/gem/.openclaw/workspace/financial-templates"

    print("Generating Additional Financial Templates...")
    print("=" * 70)

    additional_templates = [
        # More Trackers
        ("Gift_Tracker.xlsx", "Gift Tracker",
         ["Date", "Recipient", "Occasion", "Gift", "Cost", "Budget", "Notes"]),
        ("Event_Tracker.xlsx", "Event Tracker",
         ["Event", "Date", "Location", "Budget", "Actual", "Attendees", "Notes"]),
        ("Conference_Tracker.xlsx", "Conference Tracker",
         ["Conference", "Date", "Location", "Cost", "Budget", "Attended", "Notes"]),
        ("Software_License_Tracker.xlsx", "Software License Tracker",
         ["Software", "License Key", "Seats", "Cost", "Renewal Date", "Contact", "Notes"]),
        ("Domain_Tracker.xlsx", "Domain Tracker",
         ["Domain", "Registrar", "Purchase Date", "Expiration Date", "Cost", "Auto-Renew", "Notes"]),
        ("Hosting_Tracker.xlsx", "Hosting Tracker",
         ["Service", "Provider", "Cost", "Billing Cycle", "Renewal Date", "Notes"]),
        ("Utility_Tracker.xlsx", "Utility Tracker",
         ["Utility", "Provider", "Account #", "Average Cost", "Budget", "Notes"]),
        ("Subscription_Cancel_Tracker.xlsx", "Subscription Cancel Tracker",
         ["Service", "Cancel Date", "Reason", "Refund Amount", "Status", "Notes"]),

        # More Calculators
        ("Hourly_Wage_Calculator.xlsx", "Hourly Wage Calculator",
         ["Annual Salary", "Hours/Week", "Weeks/Year", "Hourly Rate", "Daily Rate", "Weekly Rate"]),
        ("Salary_Calculator.xlsx", "Salary Calculator",
         ["Hourly Rate", "Hours/Week", "Weeks/Year", "Weekly Pay", "Monthly Pay", "Annual Pay"]),
        ("Overtime_Calculator.xlsx", "Overtime Calculator",
         ["Regular Hours", "Overtime Hours", "Hourly Rate", "Overtime Rate", "Regular Pay", "Overtime Pay", "Total"]),
        ("Bonus_Calculator.xlsx", "Bonus Calculator",
         ["Base Salary", "Bonus %", "Bonus Amount", "Total Pay", "Tax Withheld", "Net Pay"]),
        ("Commission_Calculator.xlsx", "Commission Calculator",
         ["Sales Amount", "Commission Rate", "Commission", "Base Pay", "Total Pay"]),
        ("Tip_Calculator.xlsx", "Tip Calculator",
         ["Bill Amount", "Tip %", "Tip Amount", "Total", "Split Between", "Per Person"]),
        ("Discount_Calculator.xlsx", "Discount Calculator",
         ["Original Price", "Discount %", "Discount Amount", "Sale Price", "Savings"]),
        ("Tax_Calculator.xlsx", "Tax Calculator",
         ["Amount", "Tax Rate", "Tax Amount", "Total"]),
        ("Currency_Converter.xlsx", "Currency Converter",
         ["Amount", "From Currency", "To Currency", "Exchange Rate", "Converted Amount"]),
        ("Inflation_Calculator.xlsx", "Inflation Calculator",
         ["Amount", "Inflation Rate", "Years", "Future Amount", "Loss in Value"]),

        # More Lists
        ("Contact_List.xlsx", "Contact List",
         ["Name", "Company", "Phone", "Email", "Address", "Category", "Notes"]),
        ("Vendor_List.xlsx", "Vendor List",
         ["Vendor", "Contact", "Phone", "Email", "Address", "Terms", "Notes"]),
        ("Customer_List.xlsx", "Customer List",
         ["Customer", "Contact", "Phone", "Email", "Address", "Credit Limit", "Notes"]),
        ("Employee_List.xlsx", "Employee List",
         ["Employee ID", "Name", "Department", "Title", "Phone", "Email", "Hire Date", "Notes"]),
        ("Product_List.xlsx", "Product List",
         ["Product ID", "Name", "Category", "SKU", "Price", "Cost", "Supplier", "Notes"]),
        ("Service_List.xlsx", "Service List",
         ["Service ID", "Name", "Category", "Hourly Rate", "Description", "Notes"]),
        ("Account_List.xlsx", "Account List",
         ["Account #", "Account Name", "Type", "Bank", "Balance", "Notes"]),
        ("Credit_Card_List.xlsx", "Credit Card List",
         ["Card Name", "Bank", "Last 4 Digits", "Limit", "Balance", "APR", "Due Date", "Notes"]),

        # More Registers
        ("Cash_Register.xlsx", "Cash Register",
         ["Date", "Opening Balance", "Sales", "Refunds", "Payouts", "Closing Balance", "Variance"]),
        ("Petty_Cash_Register.xlsx", "Petty Cash Register",
         ["Date", "Opening Balance", "Deposits", "Withdrawals", "Closing Balance", "Notes"]),
        ("Safe_Register.xlsx", "Safe Register",
         ["Date", "Opening Balance", "Deposits", "Withdrawals", "Closing Balance", "Notes"]),

        # More Logs
        ("Activity_Log.xlsx", "Activity Log",
         ["Date", "Activity", "Description", "Duration", "Cost", "Notes"]),
        ("Change_Log.xlsx", "Change Log",
         ["Date", "Item", "Change Type", "Old Value", "New Value", "Changed By", "Notes"]),
        ("Error_Log.xlsx", "Error Log",
         ["Date", "Error Type", "Description", "Severity", "Impact", "Resolved", "Notes"]),
        ("Issue_Log.xlsx", "Issue Log",
         ["Date", "Issue", "Description", "Priority", "Assigned To", "Status", "Resolved Date", "Notes"]),

        # More Schedules
        ("Meeting_Schedule.xlsx", "Meeting Schedule",
         ["Date", "Time", "Meeting", "Attendees", "Location", "Duration", "Notes"]),
        ("Appointment_Schedule.xlsx", "Appointment Schedule",
         ["Date", "Time", "Appointment", "Contact", "Location", "Duration", "Notes"]),
        ("Task_Schedule.xlsx", "Task Schedule",
         ["Task", "Start Date", "Due Date", "Assigned To", "Priority", "Status", "Notes"]),
        ("Payment_Schedule.xlsx", "Payment Schedule",
         ["Due Date", "Payee", "Amount", "Category", "Status", "Paid Date", "Notes"]),
        ("Bill_Schedule.xlsx", "Bill Schedule",
         ["Bill", "Due Date", "Amount", "Frequency", "Auto-Pay", "Account", "Notes"]),

        # More Reports
        ("Monthly_Report.xlsx", "Monthly Report",
         ["Month", "Income", "Expenses", "Net Income", "Savings Rate", "Notes"]),
        ("Quarterly_Report.xlsx", "Quarterly Report",
         ["Quarter", "Income", "Expenses", "Net Income", "Savings Rate", "Notes"]),
        ("Annual_Report.xlsx", "Annual Report",
         ["Year", "Total Income", "Total Expenses", "Net Income", "Savings Rate", "Notes"]),
        ("Category_Report.xlsx", "Category Report",
         ["Category", "Budget", "Actual", "Variance", "Variance %", "Notes"]),
        ("Department_Report.xlsx", "Department Report",
         ["Department", "Budget", "Actual", "Variance", "Variance %", "Notes"]),
        ("Project_Report.xlsx", "Project Report",
         ["Project", "Budget", "Actual", "Variance", "Variance %", "Status", "Notes"]),

        # More Planning
        ("Financial_Goals.xlsx", "Financial Goals",
         ["Goal", "Target Amount", "Current Amount", "Target Date", "Monthly Contribution", "Status", "Notes"]),
        ("Savings_Goals.xlsx", "Savings Goals",
         ["Goal", "Target Amount", "Saved So Far", "Target Date", "Monthly Deposit", "Status", "Notes"]),
        ("Debt_Payoff.xlsx", "Debt Payoff",
         ["Creditor", "Balance", "Interest Rate", "Monthly Payment", "Payoff Date", "Priority", "Notes"]),
        ("Emergency_Fund.xlsx", "Emergency Fund",
         ["Category", "Target Amount", "Current Amount", "Monthly Contribution", "Status", "Notes"]),
        ("Retirement_Planning.xlsx", "Retirement Planning",
         ["Current Age", "Retirement Age", "Current Savings", "Monthly Contribution", "Expected Return", "Target Amount", "Notes"]),

        # More Analysis
        ("Trend_Analysis.xlsx", "Trend Analysis",
         ["Period", "Revenue", "Expenses", "Net Income", "Growth %", "Notes"]),
        ("Variance_Analysis.xlsx", "Variance Analysis",
         ["Account", "Budget", "Actual", "Variance", "Variance %", "Favorable/Unfavorable", "Notes"]),
        ("Ratio_Analysis.xlsx", "Ratio Analysis",
         ["Ratio", "Current Period", "Previous Period", "Change", "Industry Average", "Notes"]),
        ("Scenario_Analysis.xlsx", "Scenario Analysis",
         ["Metric", "Best Case", "Base Case", "Worst Case", "Notes"]),
        ("Sensitivity_Analysis.xlsx", "Sensitivity Analysis",
         ["Variable", "Low", "Base", "High", "Impact", "Notes"]),

        # More Specialized
        ("Freelancer_Tracker.xlsx", "Freelancer Tracker",
         ["Client", "Project", "Hours", "Rate", "Amount", "Invoice #", "Paid", "Notes"]),
        ("Consulting_Tracker.xlsx", "Consulting Tracker",
         ["Client", "Project", "Hours", "Rate", "Expenses", "Total", "Invoice #", "Notes"]),
        ("Contractor_Tracker.xlsx", "Contractor Tracker",
         ["Contractor", "Service", "Rate", "Hours", "Total", "Paid", "Notes"]),
        ("Gig_Tracker.xlsx", "Gig Tracker",
         ["Platform", "Gig", "Date", "Earnings", "Expenses", "Net", "Notes"]),

        # More Personal Finance
        ("Net_Worth.xlsx", "Net Worth",
         ["Asset/Liability", "Category", "Description", "Value", "Date", "Notes"]),
        ("Debt_to_Income.xlsx", "Debt to Income",
         ["Month", "Total Income", "Total Debt Payments", "DTI Ratio", "Status", "Notes"]),
        ("Expense_Ratio.xlsx", "Expense Ratio",
         ["Category", "Amount", "% of Income", "Budget %", "Variance", "Notes"]),
        ("Savings_Rate.xlsx", "Savings Rate",
         ["Month", "Income", "Expenses", "Savings", "Savings Rate", "Target Rate", "Notes"]),
    ]

    for filename, title, columns in additional_templates:
        filepath = os.path.join(output_dir, filename)
        wb = create_simple_tracker_template(filename, title, columns)
        wb.save(filepath)
        print(f"✓ {filename}")

    print("=" * 70)
    print(f"\nAdditional templates created: {len(additional_templates)}")

if __name__ == "__main__":
    generate_additional_templates()
